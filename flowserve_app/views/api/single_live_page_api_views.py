import json
from pymodbus.client import ModbusTcpClient
from django.db import connection, transaction, IntegrityError
from django.http import JsonResponse
from django.shortcuts import redirect
from django.contrib import messages
from datetime import datetime
import time
# from flowserve_app.decorators import permission_required
from django.views.decorators.csrf import csrf_exempt
from pymodbus.client import ModbusTcpClient
import pyodbc
import threading, os
from openpyxl import Workbook
from io import BytesIO
import struct
import asyncio
from flowserve_app.views.api.configuration_api_views import TestleadSmartsyncx, ABRSDatabase
from flowserve_app.services.abrs_service import ABRSService
from django.template.loader import render_to_string
from weasyprint import HTML
import matplotlib
matplotlib.use('Agg')  # Use non-GUI backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import base64

from flowserve_app.src import HmiAddress
from flowserve_app.services.single_live_page_services import(
    clear_station_1,
    clear_station_2,
    save_test_pressure_station1,
    save_test_pressure_station2,
    clear_all_livedata
    )


def check_abrs_hmi_connection(request):
    # ---- ABRS DB CHECK ----
    try:
        db = ABRSDatabase()
        abrs_status = db.test_connection()
        print("abrs status", abrs_status)
    except Exception as e:
        print("ABRS connection error:", e)
        abrs_status = False

    # ---- HMI CHECK ----
    try:
        hmi_value = getstatus(HmiAddress.VALVE_SIZE)
        hmi_status = hmi_value is not None
        print("hmi status", hmi_status)
    except Exception as e:
        print("HMI error:", e)
        hmi_status = False

    return JsonResponse({
        "abrs_connected": abrs_status,
        "hmi_connected": hmi_status,
     
    })


def getstatus(num):
    if TestleadSmartsyncx is None:
        print("[Error] HMI connection not established.")
        return None
    try:
        return TestleadSmartsyncx.read_holding_registers(num, 1).registers[0]
    except Exception as e:
        print(f"[Error] Failed to read from register {num}: {e}")
        return None
   
   
def write_to_hmi(place, value):
    try:
        TestleadSmartsyncx.write_register(place, value)
        return True
    except Exception as e:
        print(f"[Error writing to HMI] {e}")
        isconnected = False
        return False


def get_result(request):
    """Get test result from HMI - accessible globally"""
    try:
        # Read test result from HMI address 2017 for Station 1
        test_result = getstatus(HmiAddress.GOOD_NOGOOD)
        
        if test_result is not None:
            # Convert HMI result to PASS/FAIL
            result_text = "PASS" if test_result == 1 else "FAIL"
            
            return JsonResponse({
                "status": "success",
                "test_result": test_result,
                "result_text": result_text,
                "hmi_address": HmiAddress.GOOD_NOGOOD
            })
        else:
            return JsonResponse({
                "status": "error",
                "message": "Failed to read test result from HMI",
                "test_result": None,
                "result_text": "UNKNOWN"
            })
            
    except Exception as e:
        print(f"Error reading HMI test result: {e}")
        return JsonResponse({
            "status": "error",
            "message": str(e),
            "test_result": None,
            "result_text": "ERROR"
        })
   

def check_status(request):
    try:
        with connection.cursor() as cursor:

            cursor.execute("SELECT STATION_STATUS FROM master_temp_data WHERE id=1")
            row = cursor.fetchone()
            station1_status = row[0].lower() if row else "disabled"
         
            # cursor.execute("SELECT STATION_STATUS FROM master_temp_data WHERE id=2")
            # station2_status = cursor.fetchone()[0].lower()

            # s1_test_mode = getstatus(HmiAddress.MACHINE_MODE)  # HMI reading commented out
            s1_test_mode = 1  # Default to Manual mode

            # s2_test_mode = getstatus(HmiAddress.S2_MACHINE_MODE)  # HMI reading commented out
         

        # Final consistent return
        return JsonResponse({
            "status": "success",
            "station1_status": station1_status,
            # "station2_status": station2_status,
            "s1_test_mode": s1_test_mode
            # "s2_test_mode": s2_test_mode
        })

    except Exception as e:
        print("Error in check_status:", e)
        return JsonResponse({"status": "failure"})
   
   
def enabled_test_buttons(request):

    try:
        with connection.cursor() as cursor:
            # Fetch statuses
            cursor.execute("SELECT TEST_ID, VALVE_SERIAL_NO, TEST_NAME, TESTING_PR_UNIT, TESTING_DUR_UNIT FROM temp_testing_data_s1")
            s1_enabled_buttons = cursor.fetchall()

            s1_enabled_tst_buttons = []
            for btns in s1_enabled_buttons:
                s1_enabled_tst_buttons.append({
                    "id": btns[0],
                    "valve_serial_no":btns[1],
                    "name": btns[2],
                    "psr_unit":btns[3],
                    "dur_unit":btns[4]
                })

            print(s1_enabled_tst_buttons)

            # cursor.execute("SELECT TEST_ID, VALVE_SERIAL_NO, TEST_NAME, TESTING_PR_UNIT, TESTING_DUR_UNIT   FROM temp_testing_data_s2")
            # s2_enabled_test_buttons = cursor.fetchall()
           
            # s2_enabled_tst_buttons = []
            # for btns in s2_enabled_test_buttons:
            #     s2_enabled_tst_buttons.append({
            #         "id": btns[0],
            #         "valve_serial_no":btns[1],
            #         "name": btns[2],
            #         "psr_unit":btns[3],
            #         "dur_unit":btns[4]
            #     })
            # print(s2_enabled_tst_buttons)

        return JsonResponse({
            "status": "success",
            "s1_enabled_tst_btns":s1_enabled_tst_buttons
            # "s2_enabled_tst_btns":s2_enabled_tst_buttons
            })
    except:
        return JsonResponse({"status": "failure"})
   


# def getStation_values(request):
#     try:
#         with connection.cursor() as cursor:
#             cursor.execute("""
#                 SELECT VALVE_SER_NO, SIZE_NAME, CLASS_NAME, PRESSURE_UNIT, SHELL_MATERIAL_NAME, COL7_VALUE, COL8_VALUE, STATION_STATUS
#                 FROM master_temp_data
#                 WHERE id = 1
#             """)
#             S1_row = cursor.fetchone()

#             #hmi read
#             s1_open_torque = getstatus(HmiAddress.SET_OPEN_TORQUE)
#             s1_close_torque = getstatus(HmiAddress.SET_CLOSE_TORQUE)
         
#             if S1_row:
#                 valve_ser_no, size, cls, psr_unit, body_material, assembledby, testedby, station_status = S1_row

             
#                 S1_data = {
#                     "valve_ser_no":valve_ser_no,
#                     "size": size,
#                     "class": cls,
#                     "psr_unit":psr_unit,
#                     "body": body_material,
#                     "testedby": testedby,
#                     "assembledby": assembledby,
#                     "s1_torque_value":{
#                     "open_torque": s1_open_torque,
#                     "close_torque":s1_close_torque
#                     }
#                 }
               
#             cursor.execute("""
#             SELECT SIZE_ID, SIZE_NAME  
#             FROM valvesize
#             WHERE SIZE_NAME = %s
#             """, [size])
#             s1_v_size = cursor.fetchone()  
           
#             #hmi write
#             if s1_v_size:
#                 s1_size_id = s1_v_size[0]    
#                 write_to_hmi(HmiAddress.VALVE_SIZE, s1_size_id)

#             write_to_hmi(HmiAddress.VALVE_CLASS, cls)
#             write_to_hmi(HmiAddress.PRESSURE_UNIT, psr_unit)
#             if station_status == "Enabled":
#                 write_to_hmi(HmiAddress.E_D_STATUS, 1)
#             else:
#                 write_to_hmi(HmiAddress.E_D_STATUS, 0)
   
           
#             cursor.execute("""
#                 SELECT VALVE_SER_NO, SIZE_NAME, CLASS_NAME,PRESSURE_UNIT, SHELL_MATERIAL_NAME, COL7_VALUE, COL8_VALUE, STATION_STATUS
#                 FROM master_temp_data
#                 WHERE id = 2
#                 """)
#             S2_row = cursor.fetchone()

#             s2_open_torque = getstatus(HmiAddress.S2_SET_OPEN_TORQUE)
#             s2_close_torque = getstatus(HmiAddress.S2_SET_CLOSE_TORQUE)

#             if S2_row:
#                 # Pythonic tuple unpacking
#                 valve_ser_no, size, cls, psr_unit, body_material, assembledby, testedby, station_status2 = S2_row

#                 S2_data = {
#                     "valve_ser_no":valve_ser_no,
#                     "size": size,
#                     "class": cls,
#                     "psr_unit":psr_unit,
#                     "body": body_material,
#                     "testedby": testedby,
#                     "assembledby": assembledby,
#                     "s2_torque_value":{
#                     "open_torque": s2_open_torque,
#                     "close_torque":s2_close_torque
#                     }
#                 }

#                 cursor.execute("""
#                     SELECT SIZE_ID, SIZE_NAME  
#                     FROM valvesize
#                     WHERE SIZE_NAME = %s
#                     """, [size])
#                 s2_v_size = cursor.fetchone()   # get single row

#                 if s2_v_size:
#                     s2_size_id = s2_v_size[0]      # extract SIZE_ID
#                     write_to_hmi(HmiAddress.S2_VALVE_SIZE, s2_size_id)

#                 write_to_hmi(HmiAddress.S2_VALVE_CALSS, cls)
#                 write_to_hmi(HmiAddress.S2_PRESSURE_UNIT, psr_unit)
               
#                 if station_status2 == "Enabled":
#                     write_to_hmi(HmiAddress.S2_E_D_STATUS, 1)
#                 else:
#                     write_to_hmi(HmiAddress.S2_E_D_STATUS, 0)

#                 return JsonResponse({
#                     "status": "success",
#                     "S1_data": S1_data,
#                     "S2_data": S2_data
#                 })
#             else:
#                 return JsonResponse({
#                     "status": "failure",
#                     "message": "No data found for Station 1."
#                 })

#     except Exception as e:
#         print("Error in getStation1_values:", e)
#         return JsonResponse({
#             "status": "failure",
#             "message": str(e)
#         })
   


def getStation_values(request, stationNum):
    print("get values for station", stationNum)

    if stationNum not in [1, 2]:
        return JsonResponse({"status": "error", "message": "Invalid station"})

    if stationNum == 1:
        data = getStation_values1()
        print("station 1 values", data)
        return JsonResponse({"status": "success","S1_data": data})
       
    # else:
    #     data = getStation_values2()
    #     print("station2 values", data)
    #     return JsonResponse({"status": "success", "S2_data": data})


def getStation_values1():

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT VALVE_SER_NO, SIZE_NAME, CLASS_NAME, PRESSURE_UNIT, SHELL_MATERIAL_NAME, COL7_VALUE, COL8_VALUE, STATION_STATUS, TYPE_NAME
                FROM master_temp_data
                WHERE id = 1
            """)
            S1_row = cursor.fetchone()

            #hmi read
            s1_open_degree = getstatus(HmiAddress.SET_OPEN_DEGREE)
            s1_close_degree = getstatus(HmiAddress.SET_CLOSE_DEGREE)
            s1_open_torque = getstatus(HmiAddress.SET_OPEN_TORQUE)
            s1_close_torque = getstatus(HmiAddress.SET_CLOSE_TORQUE)

            # Ensure degree values are written to HMI if they exist
            if s1_open_degree is not None:
                write_to_hmi(HmiAddress.SET_OPEN_DEGREE, s1_open_degree)
            if s1_close_degree is not None:
                write_to_hmi(HmiAddress.SET_CLOSE_DEGREE, s1_close_degree)

       
            if S1_row:
                valve_ser_no, size, cls, psr_unit, body_material, assembledby, testedby, station_status, type_name = S1_row

                # Get the valve status from current_status_station1 table
                cursor.execute("""
                    SELECT RESULT
                    FROM current_status_station1
                    WHERE VALVE_SERIAL_NO = %s
                    ORDER BY DATE_TIME DESC
                    LIMIT 1
                """, [valve_ser_no])
                result_row = cursor.fetchone()
                
                valve_status = None
                if result_row:
                    result_value = result_row[0]
                    # Map result value to status text and color according to the defined mapping
                    if result_value == "0":
                        valve_status = {"text": "LEAK OBSERVED", "color": "red"}
                    elif result_value == "1":
                        valve_status = {"text": "NO LEAK OBSERVED", "color": "green"}
                    elif result_value == "2":
                        valve_status = {"text": "RUNNING", "color": "yellow"}
                    elif result_value == "3":
                        valve_status = {"text": "READY TO TEST", "color": "blue"}
                    else:
                        valve_status = {"text": result_value or "N/A", "color": "yellow"}
                else:
                    # If no result found, set default status
                    valve_status = {"text": "N/A", "color": "yellow"}

                S1_data = {
                    "valve_ser_no":valve_ser_no,
                    "size": size,
                    "class": cls,
                    "psr_unit":psr_unit,
                    "body": body_material,
                    "testedby": testedby,
                    "assembledby": assembledby,
                    "type_name": type_name,
                    "valve_status": valve_status,
                    "s1_degree_value":{
                    "open_degree": s1_open_degree,
                    "close_degree":s1_close_degree
                    },
                    "s1_torque_value":{
                    "open_torque": s1_open_torque,
                    "close_torque":s1_close_torque
                    }
                }
               
            cursor.execute("""
            SELECT SIZE_ID, SIZE_NAME  
            FROM valvesize
            WHERE SIZE_NAME = %s
            """, [size])
            s1_v_size = cursor.fetchone()  
           
            #hmi write
            if s1_v_size:
                s1_size_id = s1_v_size[0]    
                write_to_hmi(HmiAddress.VALVE_SIZE, s1_size_id)

            write_to_hmi(HmiAddress.VALVE_CLASS, cls)
            write_to_hmi(HmiAddress.PRESSURE_UNIT, psr_unit)
            write_to_hmi(HmiAddress.E_D_STATUS, 1 if station_status == "Enabled" else 0)
           
            return S1_data
           
    except Exception as e:
        print("Error in getStation1_values:", e)
        return {str(e)}


# def getStation_values2():

#     try:
#         with connection.cursor() as cursor:  
#             cursor.execute("""
#                 SELECT VALVE_SER_NO, SIZE_NAME, CLASS_NAME,PRESSURE_UNIT, SHELL_MATERIAL_NAME, COL7_VALUE, COL8_VALUE, STATION_STATUS
#                 FROM master_temp_data
#                 WHERE id = 2
#                 """)
#             S2_row = cursor.fetchone()

#             s2_open_torque = getstatus(HmiAddress.S2_SET_OPEN_TORQUE)
#             s2_close_torque = getstatus(HmiAddress.S2_SET_CLOSE_TORQUE)

#             if S2_row:
#                 # Pythonic tuple unpacking
#                 valve_ser_no, size, cls, psr_unit, body_material, assembledby, testedby, station_status2 = S2_row

#                 S2_data = {
#                     "valve_ser_no":valve_ser_no,
#                     "size": size,
#                     "class": cls,
#                     "psr_unit":psr_unit,
#                     "body": body_material,
#                     "testedby": testedby,
#                     "assembledby": assembledby,
#                     "s2_torque_value":{
#                     "open_torque": s2_open_torque,
#                     "close_torque":s2_close_torque
#                     }
#                 }

#                 cursor.execute("""
#                     SELECT SIZE_ID, SIZE_NAME  
#                     FROM valvesize
#                     WHERE SIZE_NAME = %s
#                     """, [size])
#                 s2_v_size = cursor.fetchone()   # get single row

#                 if s2_v_size:
#                     s2_size_id = s2_v_size[0]      # extract SIZE_ID
#                     write_to_hmi(HmiAddress.S2_VALVE_SIZE, s2_size_id)

#                 write_to_hmi(HmiAddress.S2_VALVE_CALSS, cls)
#                 write_to_hmi(HmiAddress.S2_PRESSURE_UNIT, psr_unit)
#                 write_to_hmi(HmiAddress.S2_E_D_STATUS, 1 if station_status2 == "Enabled" else 0)

#                 return S2_data

#     except Exception as e:
#         print("Error in getStation2_values:", e)
#         return{str(e)}
   
   
   

def get_test_set_pressure(request, id, valve_serial_no, name, stationNum, units):
   

    if request.method != "GET":
        return JsonResponse ({"error":"Invalid method"}, status = 405)
   
    print(id, valve_serial_no, name, stationNum, units)
    
    # Update Button_status to True when test button is clicked
    # try:
    #     with connection.cursor() as cursor:
    #         cursor.execute("SELECT Serial_No FROM serial_tbl WHERE Serial_No = %s", [valve_serial_no])
    #         row = cursor.fetchone()
    #         if row:
    #             cursor.execute("UPDATE serial_tbl SET Button_status = %s WHERE Serial_No = %s", [True, valve_serial_no])
    #         else:
    #             cursor.execute("INSERT INTO serial_tbl (Serial_No, Count_No, Button_status) VALUES (%s, %s, %s)", [valve_serial_no, 0, True])
    #         print(f"[TEST_BUTTON_CLICKED] Button_status set to True for {valve_serial_no}")
    # except Exception as e:
    #     print(f"[TEST_BUTTON_CLICKED] Error updating Button_status: {e}")
     
    allowed_units = ["BAR", "PSI", "KG"]

    if units not in allowed_units:
        raise ValueError("Invalid unit type")
   
    column_name = f"TESTING_PR_{units}"

    try:
        if (stationNum == 1):

            set_bubble_count = getstatus(HmiAddress.SET_BUBBLE_COUNT)
            set_clampping_psr = getstatus(HmiAddress.BUBBLE_COUNTER)
           
            with connection.cursor() as cursor:
                query = f"""
                        SELECT TEST_ID, TEST_NAME, TEST_MEDIUM,
                        TEST_CATEGORY, TESTING_PR_UNIT, {column_name}, TESTING_DUR_UNIT, TESTING_DUR_SEC
                        FROM temp_testing_data_s1
                        WHERE TEST_ID = %s AND TEST_NAME = %s
                    """
                cursor.execute(query, [id, name])
                station_1 = cursor.fetchall()

                station_data1 = {}

                for row in station_1:

                    test_id = row[0]
                    test_name = row[1]
                    pressure_unit = row[4]
                    set_pressure = row[5]
                    dur_unit = row[6]
                    set_duration = row[7]

                    station_data1= {
                    "TEST_ID": test_id ,
                    "TEST_NAME": test_name,
                    "TESTING_PSR_UNIT":pressure_unit,
                    "TESTING_PRESSURE": set_pressure,
                    "TESTING_DUR_UNIT":dur_unit,
                    "TESTING_DUR": set_duration,
                    "set_clampping_psr": set_clampping_psr,
                    "set_bubble_count": set_bubble_count
                    }

                query3 = f"""
                    SELECT CLASS_NAME
                    FROM master_temp_data
                    WHERE VALVE_SER_NO = %s
                """
                cursor.execute(query3, [valve_serial_no])
                station1_cls = cursor.fetchone()
                print(station1_cls)

               
               
                cursor.execute("""
                SELECT CLASS_ID, CLASS_NAME  
                FROM valveclass
                WHERE CLASS_NAME = %s
                """, [station1_cls])
                s1_v_class = cursor.fetchone()  
                print("class id", s1_v_class)
           
                #hmi write
                if s1_v_class:
                    s1_class = s1_v_class[0]    
                   
                    # Multiply pressure by 10 if unit is bar
                    hmi_pressure = int(set_pressure)
                    if pressure_unit.lower() == 'bar':
                        hmi_pressure = int(set_pressure * 10)

                    write_to_hmi(HmiAddress.SET_PRESSURE, hmi_pressure)
                    write_to_hmi(HmiAddress.SET_TIME, int(set_duration))
                    write_to_hmi(HmiAddress.VALVE_CLASS, int(s1_class))
                 
                    if  pressure_unit == 'psi':
                        write_to_hmi(HmiAddress.PRESSURE_UNIT, 1)

                    elif pressure_unit == 'bar':
                        write_to_hmi(HmiAddress.PRESSURE_UNIT, 2)

                    else:
                        pressure_unit == 'kg/cm2g'
                        write_to_hmi(HmiAddress.PRESSURE_UNIT, 3)

                   
                    write_to_hmi(HmiAddress.TEST_TYPE, int(test_id))
                    start_station_threads(station1_enabled = True)
                    print("station-1 thread is called for store station-1 pressure")

                    master_query = f"""
                            SELECT STANDARD_NAME, SIZE_NAME, TYPE_NAME, CLASS_NAME, SHELL_MATERIAL_NAME,
                            COL1_NAME, COL1_VALUE,
                            COL2_NAME, COL2_VALUE,
                            COL3_NAME, COL3_VALUE,
                            COL4_NAME, COL4_VALUE,
                            COL5_NAME, COL5_VALUE,
                            COL6_NAME, COL6_VALUE,
                            COL7_NAME, COL7_VALUE,
                            COL8_NAME, COL8_VALUE,
                            COL9_NAME, COL9_VALUE,
                            COL10_NAME, COL10_VALUE,
                            COL11_NAME, COL11_VALUE,
                            COL12_NAME, COL12_VALUE,
                            COL13_NAME, COL13_VALUE,
                            COL14_NAME, COL14_VALUE,
                            COL15_NAME, COL15_VALUE,
                            COL16_NAME, COL16_VALUE,
                            COL17_NAME, COL17_VALUE,
                            COL18_NAME, COL18_VALUE,
                            COL19_NAME, COL19_VALUE,
                            COL20_NAME, COL20_VALUE,
                            COL21_NAME, COL21_VALUE,
                            COL22_NAME, COL22_VALUE,
                            COL23_NAME, COL23_VALUE,
                            COL24_NAME, COL24_VALUE

                            FROM master_temp_data
                            WHERE VALVE_SER_NO = %s
                    """
                cursor.execute(master_query, [valve_serial_no])
                master_data = cursor.fetchone()
             

                if not master_data:
                    raise ValueError("No master data found for given VALVE_SER_NO")
               
                columns = [col[0] for col in cursor.description]
                data = dict(zip(columns, master_data))
             
                parameters = {}

                for i in range(1, 24):
                    col_name = data.get(f"COL{i}_NAME")
                    value = data.get(f"COL{i}_VALUE")

                    if name:  # ignore empty/null columns
                        parameters[col_name] = value
                        # print("parameters", parameters)
                       

                final_data_1 = {
                    "STANDARD_NAME": data["STANDARD_NAME"],
                    "SIZE_NAME": data["SIZE_NAME"],
                    "TYPE_NAME": data["TYPE_NAME"],
                    "CLASS_NAME": data["CLASS_NAME"],
                    "SHELL_MATERIAL_NAME": data["SHELL_MATERIAL_NAME"],
                    "PARAMETERS": parameters,
                }

                save_test_pressure_station1(id, name, valve_serial_no, station_data1, final_data_1, cursor, status=1)
               

            return JsonResponse({
                "status": "success",
                "station": 1,
                "station1_value": station_data1
            }, safe=False)
       
        # elif stationNum == 2:

        #     s2_set_bubble_count = getstatus(HmiAddress.S2_SET_BUBBLE_COUNT)
        #     s2_set_clampping_psr = getstatus(HmiAddress.S2_BUBBLE_COUNTER)

        #     with connection.cursor() as cursor:
        #         query2 = f"""
        #                 SELECT VALVE_SERIAL_NO, TEST_ID, TEST_NAME, TEST_MEDIUM,
        #                 TEST_CATEGORY, TESTING_PR_UNIT, {column_name}, TESTING_DUR_UNIT, TESTING_DUR_SEC
        #                 FROM temp_testing_data_s2
        #                 WHERE TEST_ID = %s AND TEST_NAME = %s
        #             """
        #         cursor.execute(query2, [id, name])
        #         station_2 = cursor.fetchall()
        #         print("station 2 values in temp_testing_data", station_2)

        #         station2_data = {}
        #         # master_station_data2 = []


        #         for row in station_2:
        #             s2_valve_serial_no = row[0]
        #             s2_test_id = row[1]
        #             s2_test_name = row[2]
        #             s2_pressure_unit = row[5]
        #             s2_set_pressure = row[6]
        #             s2_dur_unit = row[7]
        #             s2_set_duration = row[8]
                   
        #             station2_data={
        #             "VALVE_SERIAL_NO":s2_valve_serial_no,
        #             "TEST_ID": s2_test_id,
        #             "TEST_NAME": s2_test_name,
        #             "TESTING_PSR_UNIT":s2_pressure_unit,
        #             "TESTING_PRESSURE": s2_set_pressure,
        #             "TESTING_DUR_UNIT":s2_dur_unit,
        #             "TESTING_DUR": s2_set_duration,
        #             "set_clampping_psr": s2_set_clampping_psr,
        #             "set_bubble_count": s2_set_bubble_count
        #         }
                   
        #         vc_query = f"""
        #             SELECT CLASS_NAME
        #             FROM master_temp_data
        #             WHERE VALVE_SER_NO = %s
        #         """
        #         cursor.execute(vc_query, [valve_serial_no])
        #         station2_cls = cursor.fetchone()

               
               
        #         cursor.execute("""
        #         SELECT CLASS_ID, CLASS_NAME  
        #         FROM valveclass
        #         WHERE CLASS_NAME = %s
        #         """, [station2_cls])
        #         s2_v_class = cursor.fetchone()  
        #         print("class id", s2_v_class)
           
        #         #hmi write
        #         if s2_v_class:
        #             s2_class = s2_v_class[0]    
                   
        #         write_to_hmi(HmiAddress.S2_SET_PRESSURE, int(s2_set_pressure))
        #         write_to_hmi(HmiAddress.S2_SET_TIME, int( s2_set_duration))
        #         write_to_hmi(HmiAddress.S2_VALVE_CALSS, int(s2_class))
               
               
        #         if  s2_pressure_unit == 'psi':
        #             write_to_hmi(HmiAddress.S2_PRESSURE_UNIT, 1)

        #         elif s2_pressure_unit == 'bar':
        #             write_to_hmi(HmiAddress.S2_PRESSURE_UNIT, 2)

        #         else:
        #             s2_pressure_unit == 'kg/cm2g'
        #             write_to_hmi(HmiAddress.S2_PRESSURE_UNIT, 3)

               
        #         write_to_hmi(HmiAddress. S2_TEST_TYPE, int(s2_test_id))
        #         start_station_threads(station2_enabled = True)
        #         print("station-2 thread is called for store station-2 pressure")
       
               
        #         master_query2 = f"""
        #                 SELECT STANDARD_NAME, SIZE_NAME, TYPE_NAME, CLASS_NAME, SHELL_MATERIAL_NAME,
        #                 COL1_NAME, COL1_VALUE,
        #                 COL2_NAME, COL2_VALUE,
        #                 COL3_NAME, COL3_VALUE,
        #                 COL4_NAME, COL4_VALUE,
        #                 COL5_NAME, COL5_VALUE,
        #                 COL6_NAME, COL6_VALUE,
        #                 COL7_NAME, COL7_VALUE,
        #                 COL8_NAME, COL8_VALUE,
        #                 COL9_NAME, COL9_VALUE,
        #                 COL10_NAME, COL10_VALUE,
        #                 COL11_NAME, COL11_VALUE,
        #                 COL12_NAME, COL12_VALUE,
        #                 COL13_NAME, COL13_VALUE,
        #                 COL14_NAME, COL14_VALUE,
        #                 COL15_NAME, COL15_VALUE,
        #                 COL16_NAME, COL16_VALUE,
        #                 COL17_NAME, COL17_VALUE,
        #                 COL18_NAME, COL18_VALUE,
        #                 COL19_NAME, COL19_VALUE,
        #                 COL20_NAME, COL20_VALUE,
        #                 COL21_NAME, COL21_VALUE,
        #                 COL22_NAME, COL22_VALUE,
        #                 COL23_NAME, COL23_VALUE,
        #                 COL24_NAME, COL24_VALUE

        #                 FROM master_temp_data
        #                 WHERE VALVE_SER_NO = %s
        #             """
        #         cursor.execute(master_query2, [valve_serial_no])
        #         master_data_2 = cursor.fetchone()
             

        #         if not master_data_2:
        #             raise ValueError("No master data found for given VALVE_SER_NO")
               
        #         columns_2 = [col[0] for col in cursor.description]
        #         data_2 = dict(zip(columns_2, master_data_2))
             
        #         parameters_2 = {}

        #         for i in range(1, 24):
        #             col_name = data_2.get(f"COL{i}_NAME")
        #             value = data_2.get(f"COL{i}_VALUE")

        #             if name:  # ignore empty/null columns
        #                 parameters_2[col_name] = value
        #                 # print("parameters", parameters)
                       

        #         final_data_2 = {
        #             "STANDARD_NAME": data_2["STANDARD_NAME"],
        #             "SIZE_NAME": data_2["SIZE_NAME"],
        #             "TYPE_NAME": data_2["TYPE_NAME"],
        #             "CLASS_NAME": data_2["CLASS_NAME"],
        #             "SHELL_MATERIAL_NAME": data_2["SHELL_MATERIAL_NAME"],
        #             "PARAMETERS": parameters_2
                 
        #         }

        #         save_test_pressure_station2(id, name, valve_serial_no, station2_data, final_data_2, cursor)
               


        #     return JsonResponse({
        #         "status": "success",
        #         "station": 2,
        #         "station2_value": station2_data
        #     }, safe=False)

    except Exception as e:
        print("ERROR:", e)
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


def station_live_values(request, stationNum, id, valve_serial_no):
    if stationNum not in [1, 2]:
        return JsonResponse({"error": "Invalid station"})

    if stationNum == 1:
        data =  get_live_pressure_data1(request, id, valve_serial_no, stationNum)

       
    # else:
    #     data = get_live_pressure_data2(request, id, valve_serial_no, stationNum)

    return JsonResponse({
        "status": "success",
        "station": stationNum,
        "data": data
    })


@csrf_exempt
def get_test_mode(request, stationNum):
    """API endpoint to get current test mode for a station"""
    try:
        if stationNum == 1:
            # test_mode = getstatus(HmiAddress.MACHINE_MODE)  # HMI reading commented out
            test_mode = 1  # Default to Manual mode
        else:
            return JsonResponse({"status": "error", "message": "Invalid station"}, status=400)
        
        return JsonResponse({
            "status": "success",
            "station": stationNum,
            "test_mode": test_mode
        })
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)



def get_history_values(request, stationNum, testId, valve_serial_no):
    if stationNum not in [1]:
        return JsonResponse({"error": "Invalid station"})

    if stationNum == 1:
        data =  get_history_prssure_data1(request, testId, valve_serial_no, stationNum)
    # else:
    #     data = get_history_prssure_data2(request, testId, valve_serial_no, stationNum)

    return JsonResponse({
        "status": "success",
        "station": stationNum,
        "data": data
    })


def get_history_prssure_data1(request, testId, valve_serial_no, stationNum):


    with connection.cursor() as cursor:

        cursor.execute("""
        SELECT PRESSURE, TIMER_STATUS, DATE_TIME, RESULT
        FROM current_status_station1
        WHERE TEST_ID = %s AND VALVE_SERIAL_NO = %s
        ORDER BY id ASC
    """, [testId, valve_serial_no])

        rows = cursor.fetchall() or []

    data = []
    for pressure, timer_status, date_time, result in rows:
        data.append({
            "pressure": float(pressure) if pressure is not None else 0.0,
            "timerStatus": timer_status,
            "time": date_time.strftime("%H:%M:%S") if date_time else "",
            # "time": date_time if date_time else "",
            "result": result
        })

    return data
   



# def get_history_prssure_data2(request, testId, valve_serial_no, stationNum):


#     with connection.cursor() as cursor:

#         cursor.execute("""
#         SELECT PRESSURE, TIMER_STATUS, DATE_TIME, RESULT
#         FROM current_status_station2
#         WHERE TEST_ID = %s AND VALVE_SERIAL_NO = %s
#         ORDER BY id ASC
#     """, [testId, valve_serial_no])

#         rows2 = cursor.fetchall() or []

#     data = []
#     for pressure, timer_status, date_time, result in rows2:
#         data.append({
#             "pressure": float(pressure) if pressure is not None else 0.0,
#             "timerStatus": timer_status,
#             # "time": date_time.strftime("%H:%M:%S") if date_time else "",
#             "time": date_time if date_time else "",
#             "result": result
#         })

#     return data




def get_live_pressure_data1(request, id, valve_serial_no, stationNum):

    actual_pre = None
    actual_timer_status = None
    actual_time = None
    result = None

    with connection.cursor() as cursor:

        # CASE 1: Initial live pressure (NO test yet)
        if int(id) == 0:
            cursor.execute("""
                SELECT PRESSURE, TIMER_STATUS, DATE_TIME, RESULT
                FROM current_status_station1
                WHERE VALVE_SERIAL_NO = %s
                ORDER BY id DESC
                LIMIT 10
            """, [valve_serial_no])

        # # CASE 2: Test running (test-specific pressure)
        else:
            cursor.execute("""
                SELECT PRESSURE, TIMER_STATUS, DATE_TIME, RESULT
                FROM current_status_station1
                WHERE TEST_ID = %s AND VALVE_SERIAL_NO = %s
                ORDER BY id DESC
                LIMIT 1
            """, [id, valve_serial_no])

        value1 = cursor.fetchone()
       
        actual_duration = getstatus(HmiAddress.ACTUAL_TIME)
        actual_open_degree = getstatus(HmiAddress.ACTUAL_OPEN_DEGREE)
        actual_close_degree =getstatus(HmiAddress.ACTUAL_CLOSE_DEGREE)
        set_open_torque = getstatus(HmiAddress.SET_OPEN_TORQUE)
        set_close_torque = getstatus(HmiAddress.SET_CLOSE_TORQUE)
        actual_open_torque = getstatus(HmiAddress.ACTUAL_OPEN_TORQUE)
        actual_close_torque = getstatus(HmiAddress.ACTUAL_CLOSE_TORQUE)
        result_value1 = getstatus(HmiAddress.GOOD_NOGOOD)
        actual_bubble = getstatus(HmiAddress.TEMPERATURE)
        allowed_bubble = getstatus(HmiAddress.SET_BUBBLE_COUNT)
        actual_clamping_psr = getstatus(HmiAddress.RESULT_PRESSURE)
        set_clamping_psr = getstatus(HmiAddress.BUBBLE_COUNTER)
        alarm_msg = getstatus(HmiAddress.ALARM_MESSAGE)
        print(f"Raw alarm status from HMI: {alarm_msg}")
        
        # Get test mode from HMI (commented out to default to Manual)
        # test_mode = getstatus(HmiAddress.MACHINE_MODE)  # HMI reading commented out
        test_mode = 1  # Default to Manual mode
        print(f"Test mode from HMI: {test_mode}")
        
        # Get alarm name from database
        alarm_name = None
        if alarm_msg is not None:
            with connection.cursor() as cursor:
                cursor.execute("select ALARM_NAME from alarm where ALARM_ID = %s", [alarm_msg])
                alarm_result = cursor.fetchone()
                alarm_name = alarm_result[0] if alarm_result else "Unknown Alarm"
                print(f"Alarm ID: {alarm_msg}, Alarm Name: {alarm_name}")
        else:
            alarm_name = "No Alarm"
            print("No alarm detected")
        # leak_pressure = getstatus(HmiAddress.)

        if value1:
            actual_pre = value1[0]
            actual_timer_status = value1[1]
            actual_time = value1[2]
            result = value1[3]

    # Map RESULT values to status text and color
    status_mapping = {
        0: {"text": "LEAK OBSERVED", "color": "red"},
        1: {"text": "NO LEAK OBSERVED", "color": "green"},
        2: {"text": "RUNNING", "color": "yellow"},
        3: {"text": "READY TO TEST", "color": "blue"}
    }
    
    # Convert result to int if it's numeric, otherwise keep as is
    result_value = result
    if result is not None:
        try:
            result_int = int(result)
            if result_int in status_mapping:
                result_value = status_mapping[result_int]
            else:
                result_value = {"text": str(result), "color": "yellow"}
        except (ValueError, TypeError):
            result_value = {"text": str(result), "color": "yellow"}
    else:
        result_value = {"text": "N/A", "color": "yellow"}
    
    return {
        "connected": True,
        "actualPressure": float(actual_pre) if actual_pre is not None else 0.0,
        "time": str(actual_time) if actual_time else "",
        "timerStatus": actual_timer_status,
        "result": result,  # Keep original result for backward compatibility
        "valve_status": result_value,  # New structured status with text and color
        "actual_duration":actual_duration,
        "actual_open_degree": actual_open_degree,
        "actual_close_degree":actual_close_degree,
        "set_open_torque": set_open_torque,
        "set_close_torque": set_close_torque,
        "actual_open_torque": actual_open_torque,
        "actual_close_torque": actual_close_torque,
        "result-value":result_value1,
        "actual_bubbles": actual_bubble,
        "allowed_bubbles": allowed_bubble,
        "actual_clamping_psr":actual_clamping_psr,
        "set_clamping_psr": set_clamping_psr,
        'alarm_msg': alarm_name,
        'test_mode': test_mode
    }


# def get_live_pressure_data2(request, id, valve_serial_no, stationNum):

#     s2_actual_pre = None
#     s2_actual_timer_status = None
#     s2_actual_time = None
#     s2_result = None

#     with connection.cursor() as cursor:

#         if int(id) == 0:
#             cursor.execute("""
#                 SELECT PRESSURE, TIMER_STATUS, DATE_TIME, RESULT
#                 FROM current_status_station2
#                 WHERE VALVE_SERIAL_NO = %s
#                 ORDER BY id DESC
#                 LIMIT 1
#             """, [valve_serial_no])
#         else:
#             cursor.execute("""
#                 SELECT PRESSURE, TIMER_STATUS, DATE_TIME, RESULT
#                 FROM current_status_station2
#                 WHERE TEST_ID = %s AND VALVE_SERIAL_NO = %s
#                 ORDER BY id DESC
#                 LIMIT 1
#             """, [id, valve_serial_no])

#         value2 = cursor.fetchone()

#         s2_actual_duration = getstatus(HmiAddress.S2_ACTUAL_TIME)
#         s2_actual_open_torque = getstatus(HmiAddress.S2_ACTUAL_OPEN_TORQUE)
#         s2_actual_close_torque =getstatus(HmiAddress.S2_ACTUAL_CLOSE_TORQUE )
#         s2_actual_bubble = getstatus(HmiAddress.S2_ACTUAL_BUBBLE_COUNT)
#         s2_actual_clamping_psr=getstatus(HmiAddress.S2_RESULT_PRESSURE)
#         result_value2 = getstatus(HmiAddress.S2_GOOD_NOGOOD)
#         # leak_pressure = getstatus(HmiAddress.)

#         if value2:
#             s2_actual_pre = value2[0]
#             s2_actual_timer_status = value2[1]
#             s2_actual_time = value2[2]
#             s2_result = value2[3]

#     return {
#         "connected": True,
#         "actualPressure": float(s2_actual_pre) if s2_actual_pre is not None else 0.0,
#         "time": str(s2_actual_time) if s2_actual_time else "",
#         "timerStatus": s2_actual_timer_status,
#         "result": s2_result,
#         "actual_duration": s2_actual_duration,
#         "actual_open_toq": s2_actual_open_torque,
#         "actual_close_toq":s2_actual_close_torque,
#         "actual_bubbles": s2_actual_bubble,
#         "actual_clamping_psr":s2_actual_clamping_psr,
#         "result-value":result_value2
#     }



station1_stop = threading.Event()
# station2_stop = threading.Event()

station1_thread = None
# station2_thread = None

def store_pressure_station1():
    print("Station-1 pressure thread started")
    
    # Track previous timer status to detect OFF transition
    previous_timer_status = {}  # {test_id: timer_status}

    while not station1_stop.is_set():
        try:
            pressure = getstatus(HmiAddress.ACTUAL_PRESSURE)
            timer_status = getstatus(HmiAddress.START_STOP)
            test_id = getstatus(HmiAddress.TEST_TYPE)

            with connection.cursor() as cursor:
                query = f"""
                        SELECT VALVE_SER_NO,
                        PRESSURE_UNIT
                        FROM master_temp_data
                        WHERE STATION_STATUS = "Enabled" and ID = 1
                    """
                cursor.execute(query)
                pressure_valveserial = cursor.fetchall()
                print("store pressure value print", pressure_valveserial)

                for row in pressure_valveserial:

                    serial_no = row[0]
                    pressure_unit = row[1].lower()
                   
                    # Read pressure based on unit
                    if  pressure_unit == 'psi':
                        pressure = getstatus(HmiAddress.ACTUAL_PRESSURE)
                    elif  pressure_unit == 'bar':
                        pressure_bar = getstatus(HmiAddress.ACTUAL_PRESSURE)
                        pressure = pressure_bar / 10
                    elif  pressure_unit == 'kg/cm2g':
                        pressure_kg = getstatus(HmiAddress.ACTUAL_PRESSURE)
                        pressure = pressure_kg / 10
                    else:
                        pressure = getstatus(HmiAddress.ACTUAL_PRESSURE)
               
                    # Read other HMI values
                    result =  getstatus(HmiAddress.GOOD_NOGOOD)
                    timer_status = getstatus(HmiAddress.START_STOP)
                    s1_test_id = getstatus(HmiAddress.TEST_TYPE)

                    query = """
                        SELECT TEST_NAME
                        FROM temp_testing_data_s1
                        WHERE TEST_ID = %s
                    """

                    cursor.execute(query, [s1_test_id])
                    s1_row = cursor.fetchone()

                    if s1_row:
                        s1_test_name =s1_row[0]
                    else:
                        s1_test_name = None

                    # Detect timer OFF transition (1 -> 0) and update test status
                    prev_status = previous_timer_status.get(s1_test_id, 0)
                    if prev_status == 1 and timer_status == 0:
                        # Timer just turned OFF, read result and update corresponding test status
                        test_result = getstatus(HmiAddress.GOOD_NOGOOD)
                        
                        # Map test ID to test status HMI address
                        test_status_map = {
                            1: HmiAddress.HYDRO_SHELL_TEST_STATUS,
                            2: HmiAddress.HYDRO_SEAT_P_TEST_STATUS,
                            3: HmiAddress.HYDRO_SEAT_N_TEST_STATUS,
                            4: HmiAddress.AIR_SEAT_P_TEST_STATUS,
                            5: HmiAddress.AIR_SEAT_N_TEST_STATUS
                        }
                        
                        if s1_test_id in test_status_map:
                            status_address = test_status_map[s1_test_id]
                            write_to_hmi(status_address, test_result)
                            print(f"[TIMER_OFF] Test ID {s1_test_id} completed with result {test_result}, updated HMI address {status_address}")
                    
                    # Update previous timer status
                    previous_timer_status[s1_test_id] = timer_status

            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO current_status_station1
                    (VALVE_SERIAL_NO, TEST_ID, TEST_NAME, PRESSURE, TIMER_STATUS, RESULT)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, [serial_no, test_id, s1_test_name, pressure, timer_status, result])

           
            print(f"Data stored for S1: Serial={serial_no}, Pressure={pressure}, Test_ID={s1_test_id}, Status={timer_status}")
            print("THREAD ID:", threading.get_ident())

        except Exception as e:
            print("S1 error:", e)

        time.sleep(0.7)


# def store_pressure_station2():
#     print("Station-2 pressure thread started")

#     while not station2_stop.is_set():
#         try:
#             s2_pressure = getstatus(HmiAddress.S2_ACTUAL_PRESSURE)
#             s2_timer_status = getstatus(HmiAddress.S2_START_STOP)

#             with connection.cursor() as cursor:
#                 query = f"""
#                         SELECT VALVE_SER_NO,
#                         PRESSURE_UNIT
#                         FROM master_temp_data
#                         WHERE STATION_STATUS = "Enabled" AND ID=2
#                     """
#                 cursor.execute(query)
#                 pressure_valveserial_2 = cursor.fetchall()
#                 print("store pressure value print", pressure_valveserial_2)


#                 for row2 in pressure_valveserial_2:

#                     s2_serial_no = row2[0]
#                     s2_pressure_unit = row2[1].lower()
#                     # print("pressure unit for live", s2_pressure_unit)
                   
#                     # Read pressure based on unit
#                     if  s2_pressure_unit == 'psi':
#                         s2_pressure = getstatus(HmiAddress.S2_ACTUAL_PRESSURE)
#                         print("psi",s2_pressure)
#                     elif  s2_pressure_unit == 'bar':
#                         pressure_bar = getstatus(HmiAddress.S2_ACTUAL_PRESSURE)
#                         s2_pressure = pressure_bar / 10
#                         print("bar",s2_pressure)
#                     elif  s2_pressure_unit == 'kg/cm2g':
#                         pressure_kg = getstatus(HmiAddress.S2_ACTUAL_PRESSURE)
#                         s2_pressure = pressure_kg / 10
#                         print("kg",s2_pressure)
#                     else:
#                         s2_pressure = getstatus(HmiAddress.S2_ACTUAL_PRESSURE)
               
#                     # Read other HMI values
#                     s2_result =  getstatus(HmiAddress.S2_GOOD_NOGOOD)
#                     s2_timer_status = getstatus(HmiAddress.S2_START_STOP)
#                     s2_test_id = getstatus(HmiAddress.S2_TEST_TYPE)

#                     query = """
#                         SELECT TEST_NAME
#                         FROM temp_testing_data_s2
#                         WHERE TEST_ID = %s
#                     """

#                     cursor.execute(query, [s2_test_id])
#                     s2_row = cursor.fetchone()

#                     if s2_row:
#                         s2_test_name = s2_row[0]
#                     else:
#                         s2_test_name = None


#             # with connection.cursor() as cursor:
#                 cursor.execute("""
#                      INSERT INTO current_status_station2
#                         (VALVE_SERIAL_NO, TEST_ID, TEST_NAME, PRESSURE, TIMER_STATUS, RESULT)
#                         VALUES (%s, %s, %s, %s, %s, %s)
#                         """,
#                         [s2_serial_no, s2_test_id, s2_test_name, s2_pressure, s2_timer_status, s2_result]
#                     )
#             print(f"Data stored for S2: Serial={s2_serial_no}, Pressure={s2_pressure}, Test_ID={s2_test_id}, Status={s2_timer_status}")
#             print("THREAD ID:", threading.get_ident())

#         except Exception as e:
#             print("S2 error:", e)

#         time.sleep(1)


def start_station_threads(station1_enabled = False, station2_enabled = False):
    global station1_thread

    # Station 1
    if station1_enabled:
        if not station1_thread or not station1_thread.is_alive():
            station1_stop.clear()
            station1_thread = threading.Thread(
                target=store_pressure_station1,
                daemon=True
            )
            station1_thread.start()
            print("Started Station-1 thread")
   

    # Station 2
    # if station2_enabled:
    #     if not station2_thread or not station2_thread.is_alive():
    #         station2_stop.clear()
    #         station2_thread = threading.Thread(
    #             target=store_pressure_station2,
    #             daemon=True
    #         )
    #         station2_thread.start()
    #         print("Started Station-2 thread")
 


def stop_station1():
    global station1_thread
    if station1_thread and station1_thread.is_alive():
        station1_stop.set()
        print("Stopping Station-1 thread")

# def stop_station2():
#     global station2_thread
#     if station2_thread and station2_thread.is_alive():
#         station2_stop.set()
#         print("Stopping Station-2 thread")


station1_enabled = getstatus(HmiAddress.E_D_STATUS) == 1
# station2_enabled = getstatus(HmiAddress.S2_E_D_STATUS) == 1

if os.environ.get("RUN_MAIN") == "true":
    # start_pressure_threads()
    start_station_threads(station1_enabled)


def save_initial_pressure(request, id, valve_serial_no, stationNum):

    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)
     
    try:
        stationNum = int(stationNum)

        with connection.cursor() as cursor:

            if stationNum == 1:
                # Get current pressure from HMI to save as Start Pressure
                start_pressure = getstatus(HmiAddress.ACTUAL_PRESSURE)
                cursor.execute("""
                    UPDATE temp_pressure_analysis
                    SET  START_PRESSURE = %s
                    WHERE TEST_ID = %s AND VALVE_SER_NO = %s AND CYCLE_COMPLETE = 'No'
                """, [start_pressure, id, valve_serial_no])
                return JsonResponse({"status": "success", "message": "Initial pressure saved", "value": start_pressure})


    except ValueError:
        return JsonResponse(
            {"error": "Station number must be integer"},
            status=400
        )

    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)

@csrf_exempt
def resetallstation(request):
    try:
        cycle_start_stop_status = getstatus(HmiAddress.CYCLE_START_STOP_STATUS)
        
        if cycle_start_stop_status == 0:
            clear_all_livedata()
            stop_station1()
            write_to_hmi(HmiAddress.TEST_TYPE, 0)
            return JsonResponse({"status":"success"})
        else:
            return JsonResponse({"status":"failure",'message':'Kindly Stop the cycle to Reset'})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@csrf_exempt
def save_final_pressure(request, testId, valve_serial_no, stationNum):

    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)

    try:
        stationNum = int(stationNum)

        body = json.loads(request.body.decode("utf-8"))
        print("received body", body)

        start_pressure = body.get("start_pressure")
        end_pressure   = body.get("end_pressure")
        start_time = body.get("start_time")
        end_time = body.get("end_time")
        result_psr = body.get("result_psr")
        actual_time = body.get("actual_dur")
        clampping_psr = body.get("clamping_psr")
        # Get torque values from HMI instead of frontend
        open_torque = getstatus(HmiAddress.ACTUAL_OPEN_TORQUE)
        close_torque = getstatus(HmiAddress.ACTUAL_CLOSE_TORQUE)
        pressure_drop = body.get("pressure_drop")
        test_result = body.get("test_result")

        # Convert test_result to status: 1 for PASS, 0 for FAIL
        # STATUS: 1 = Running, 0 = Completed (regardless of PASS/FAIL)
        status = 0  # Test is completed

        # s1_test_result = getstatus(HmiAddress.GOOD_NOGOOD)
        # print("result value from hmi", s1_test_result)

        # if s1_test_result == 1:
        #     result = "PASS"
        # else:
        #     result = "FAIL"

        if start_pressure is None or end_pressure is None:
            return JsonResponse({"error": "Missing pressure values"},status=400)
       

        with connection.cursor() as cursor:
            if stationNum == 1:
                # First table - includes STATUS field
                cursor.execute("""
                    UPDATE temp_pressure_analysis
                    SET
                        ACTUAL_PRESSURE = %s,
                        START_PRESSURE  = %s,
                        RESULT_PRESSURE = %s,
                        LEAK_PRESSURE = %s,
                        ACTUAL_TIME=%s,
                        CLAMPING_PRESSURE = %s,
                        ACTUAL_OPEN_TORQUE =%s,
                        ACTUAL_CLOSE_TORQUE = %s,
                        `START` = %s,
                        `END` = %s,
                        VALVE_STATUS = %s,
                        STATUS = %s,
                        DATE_TIME = NOW()
                    WHERE TEST_ID = %s
                    AND VALVE_SER_NO = %s
                    AND CYCLE_COMPLETE = 'No'
                """, [
                    result_psr,
                    start_pressure,
                    end_pressure,
                    pressure_drop,
                    actual_time,
                    clampping_psr,
                    open_torque,
                    close_torque,
                    start_time,
                    end_time,
                    test_result,
                    status,
                    testId,
                    valve_serial_no
                ])

                # Second table
                cursor.execute("""
                    UPDATE pressure_analysis
                    SET
                        ACTUAL_PRESSURE = %s,
                        START_PRESSURE  = %s,
                        RESULT_PRESSURE = %s,
                        LEAK_PRESSURE = %s,
                        ACTUAL_TIME=%s,
                        CLAMPING_PRESSURE = %s,
                        ACTUAL_OPEN_TORQUE =%s,
                        ACTUAL_CLOSE_TORQUE = %s,
                        `START` = %s,
                        `END` = %s,
                        VALVE_STATUS = %s ,
                        DATE_TIME = NOW()
                    WHERE TEST_ID = %s
                    AND VALVE_SER_NO = %s
                    AND CYCLE_COMPLETE = 'No'
                """, [
                    result_psr,
                    start_pressure,
                    end_pressure,
                    pressure_drop,
                    actual_time,
                    clampping_psr,
                    open_torque,
                    close_torque,
                    start_time,
                    end_time,
                    test_result,
                    testId,
                    valve_serial_no
                ])


            # elif stationNum == 2:
            #     cursor.execute("""
            #         UPDATE temp_pressure_analysis
            #         SET
            #             ACTUAL_PRESSURE = %s,
            #             START_PRESSURE = %s,
            #             RESULT_PRESSURE= %s,
            #             LEAK_PRESSURE = %s,
            #             ACTUAL_TIME=%s,
            #             CLAMPING_PRESSURE = %s,
            #             ACTUAL_OPEN_TORQUE =%s,
            #             ACTUAL_CLOSE_TORQUE = %s,
            #             `START` = %s,
            #             `END` = %s,
            #             VALVE_STATUS = %s,
            #             STATUS = %s,
            #             DATE_TIME = NOW()
            #         WHERE TEST_ID = %s
            #         AND VALVE_SER_NO = %s
            #         AND CYCLE_COMPLETE = 'No'
            #     """, [
            #         result_psr,
            #         start_pressure,
            #         end_pressure,
            #         pressure_drop,
            #         actual_time,
            #         clampping_psr,
            #         open_torque,
            #         close_torque,
            #         start_time,
            #         end_time,
            #         test_result,
            #         status,
            #         testId,
            #         valve_serial_no
            #     ])
            #     cursor.execute("""
            #         UPDATE pressure_analysis
            #         SET
            #             ACTUAL_PRESSURE = %s,
            #             START_PRESSURE  = %s,
            #             RESULT_PRESSURE = %s,
            #             LEAK_PRESSURE = %s,
            #             ACTUAL_TIME=%s,
            #             CLAMPING_PRESSURE = %s,
            #             ACTUAL_OPEN_TORQUE =%s,
            #             ACTUAL_CLOSE_TORQUE = %s,
            #             `START` = %s,
            #             `END` = %s,
            #             VALVE_STATUS = %s,
            #             DATE_TIME = NOW()
            #         WHERE TEST_ID = %s
            #         AND VALVE_SER_NO = %s
            #         AND CYCLE_COMPLETE = 'No'
            #     """, [
            #         result_psr,
            #         start_pressure,
            #         end_pressure,
            #         pressure_drop,
            #         actual_time,
            #         clampping_psr,
            #         open_torque,
            #         close_torque,
            #         start_time,
            #         end_time,
            #         test_result,
            #         testId,
            #         valve_serial_no
            #     ])
        return JsonResponse({
            "status": "success",
            "message": "Final pressure saved"
        })

    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)
   

@csrf_exempt
def get_test_result_status(request, stationNum, valve_serial_no):
    """Get test result status for button color coding"""
   
    if request.method != "GET":
        return JsonResponse({"error": "Invalid method"}, status=405)
   
    try:
        stationNum = int(stationNum)
       
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT TEST_ID, STATUS, VALVE_STATUS, START, END
                FROM temp_pressure_analysis
                WHERE VALVE_SER_NO = %s
                AND CYCLE_COMPLETE = 'No'
            """, [valve_serial_no])
           
            rows = cursor.fetchall()
           
            test_statuses = {}
            for row in rows:
                test_id = row[0]
                status = row[1]  # 1 = Running, 0 = Completed
                valve_status = row[2]  # "PASS" or "FAIL"
                start_time = row[3]
                end_time = row[4]
               
                # Determine if test is running (STATUS = 1)
                running = status == 1
                # Determine if test is completed (has VALVE_STATUS)
                completed = valve_status is not None
               
                test_statuses[test_id] = {
                    "status": status,
                    "valve_status": valve_status,
                    "running": running,
                    "completed": completed
                }
           
            return JsonResponse({
                "status": "success",
                "station": stationNum,
                "test_statuses": test_statuses
            })
           
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@csrf_exempt
def delete_and_retest(request, stationNum, testId, valve_serial_no):

    if request.method != "DELETE":
        return JsonResponse({"error": "Invalid method"}, status=405)

    try:
        stationNum = int(stationNum)
        testId = int(testId)

        # Choose table based on station
        if stationNum == 1:
            table = "current_status_station1"
        elif stationNum == 2:
            table = "current_status_station2"
        else:
            return JsonResponse(
                {"error": "Invalid station number"},
                status=400
            )

        with transaction.atomic():
            with connection.cursor() as cursor:
                cursor.execute(
                    f"""
                    DELETE FROM {table}
                    WHERE TEST_ID = %s
                    AND VALVE_SERIAL_NO = %s
                    """,
                    [testId, valve_serial_no]
                )
               
                cursor.execute(
                    """
                    UPDATE temp_pressure_analysis
                    SET
                        STATUS = '2',
                        VALVE_STATUS = 'No'
                    WHERE TEST_ID = %s
                    AND VALVE_SER_NO = %s
                    AND CYCLE_COMPLETE = 'No'
                    """,
                    [testId, valve_serial_no]
                )
               
       
            return JsonResponse({
                "success": True,
                "message": "Previous test data deleted successfully",
                "station": stationNum,
                "testId": testId,
                "valve_serial_no": valve_serial_no
            })

    except ValueError:
        return JsonResponse(
            {"error": "Invalid station or test ID"},
            status=400
        )

    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@csrf_exempt
def delete_incomplete_test(request,serialNo, stationNum):

    print("received parameters", stationNum, serialNo )

    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)

    try:
        stationNum = int(stationNum)

        if stationNum not in [1]:
            return JsonResponse({"error": "Invalid station"})

        if stationNum == 1:
            pressure_drain_status = getstatus(HmiAddress.PRESSURE_DRAIN) 
            cycle_start_stop_status = getstatus(HmiAddress.CYCLE_START_STOP_STATUS) 

        
        # Only allow cycle complete 
            if pressure_drain_status != 0 or cycle_start_stop_status != 0:
                return JsonResponse({"status": "failure",'message':'Kindly Ensure that Pressure Drain is Completed and Cycle is Stopped to Delete Incomplete Test'}, status=400)
            else:
                data =  delete_incomplete_test_station1(request, serialNo)

        # else:
        #     data = delete_incomplete_test_station2(request, serialNo)

        return JsonResponse({
            "success": True,  
            "status": "",
            "station":stationNum,
            "data": data
        })
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)




def delete_incomplete_test_station1(request, v_serial_no):  

        with transaction.atomic():
            with connection.cursor() as cursor:

                cursor.execute(f"""
                    UPDATE master_temp_data
                    SET STATION_STATUS = %s
                    where VALVE_SER_NO = %s
                """, ["Disabled", v_serial_no])

                cursor.execute("""
                    TRUNCATE TABLE temp_testing_data_s1
                    """)

                cursor.execute("""
                    TRUNCATE TABLE temp_pressure_analysis
                    """)
                cursor.execute(f"""      
                    DELETE FROM pressure_analysis
                    WHERE VALVE_SER_NO = %s AND
                    CYCLE_COMPLETE = "No"
                    """,
                    [v_serial_no]
                )
                cursor.execute(f"""      
                    DELETE FROM current_status_station1
                    WHERE VALVE_SERIAL_NO = %s
                    """,
                    [v_serial_no]
                )

                cursor.execute("TRUNCATE TABLE current_status_station1")
                cursor.execute("update abrs_result_status set STATUS = '0' where SERIAL_NO = %s", [v_serial_no])
                stop_station1()
                
                # for attribute_name in dir(HmiAddress):
                #     if attribute_name.startswith('S1_'):
                #         address = getattr(HmiAddress, attribute_name)
                #         write_to_hmi(address, 0)
                #         write_to_hmi(HmiAddress.PRESSURE_UNIT, 0)

                write_to_hmi(HmiAddress.TEST_TYPE, 0)
                write_to_hmi(HmiAddress.VALVE_SIZE, 0)
                write_to_hmi(HmiAddress.VALVE_CLASS, 0)
                write_to_hmi(HmiAddress.SET_PRESSURE, 0)
                write_to_hmi(HmiAddress.SET_HOLDING_TIME, 0)
                write_to_hmi(HmiAddress.SET_OPEN_DEGREE, 0)
                write_to_hmi(HmiAddress.SET_CLOSE_DEGREE, 0)
                write_to_hmi(HmiAddress.PRESSURE_UNIT, 0)
                write_to_hmi(HmiAddress.SET_TIME, 0)
                write_to_hmi(HmiAddress.GOOD_NOGOOD, 0)
                write_to_hmi(HmiAddress.E_D_STATUS, 0)
                write_to_hmi(HmiAddress.HYDRO_SHELL_E_D_STATUS, 0)
                write_to_hmi(HmiAddress.HYDRO_SEAT_P_E_D_STATUS, 0)
                write_to_hmi(HmiAddress.HYDRO_SEAT_N_E_D_STATUS, 0)
                write_to_hmi(HmiAddress.AIR_SEAT_P_E_D_STATUS, 0)
                write_to_hmi(HmiAddress.AIR_SEAT_N_E_D_STATUS, 0)

                write_to_hmi(HmiAddress.HYDRO_SHELL_TEST_STATUS, 0)
                write_to_hmi(HmiAddress.HYDRO_SEAT_P_TEST_STATUS, 0)
                write_to_hmi(HmiAddress.HYDRO_SEAT_N_TEST_STATUS, 0)
                write_to_hmi(HmiAddress.AIR_SEAT_P_TEST_STATUS, 0)
                write_to_hmi(HmiAddress.AIR_SEAT_N_TEST_STATUS, 0)

        return {
        "valve_serial_no": v_serial_no,
        "station": 1,
        "station_status": "Disabled",
        "deleted": True
        }



# def delete_incomplete_test_station2(request, v_serial_no):  

#             with transaction.atomic():
#                 with connection.cursor() as cursor:

#                     cursor.execute(f"""
#                         UPDATE master_temp_data
#                         SET STATION_STATUS = "Disabled"
#                         where VALVE_SER_NO = %s
#                     """, [v_serial_no])

#                     cursor.execute(f"""      
#                         DELETE FROM temp_testing_data_s2
#                         WHERE VALVE_SERIAL_NO = %s
#                         """,
#                         [v_serial_no]
#                     )
#                     cursor.execute(f"""      
#                         DELETE FROM temp_pressure_analysis
#                         WHERE VALVE_SER_NO = %s AND
#                         CYCLE_COMPLETE = "No"
#                         """,
#                         [v_serial_no]
#                     )
#                     cursor.execute(f"""      
#                         DELETE FROM pressure_analysis
#                         WHERE VALVE_SER_NO = %s AND
#                         CYCLE_COMPLETE = "No"
#                         """,
#                         [v_serial_no]
#                     )
#                     cursor.execute(f"""      
#                         DELETE FROM current_status_station2
#                         WHERE VALVE_SERIAL_NO = %s
#                         """,
#                         [v_serial_no]
#                     )
   
#             return {
#                 "valve_serial_no": v_serial_no,
#                 "station": 2,
#                 "station_status": "Disabled",
#                 "deleted": True
#             }

def abrs_push(valve_serial_no, testId):

    TEST_PAIR_MAPPING = {
        1: (4, 5),
        2: (6, 7),
        3: (8, 9),
        4: (10, 11),
        5: (12, 13),
    }

    COLUMN_MAPPING = {
        1:'COL1_VALUE',
        2:'COL2_VALUE',
        3:'COL3_VALUE',
        4:'COL4_VALUE',
        5:'COL5_VALUE',
        6:'COL6_VALUE',
        7:'COL7_VALUE',
        8:'COL8_VALUE',
        9:'COL9_VALUE',
        10:'COL10_VALUE',
        11:'COL11_VALUE',
        12:'COL12_VALUE',
        13:'COL13_VALUE'
    }

    with connection.cursor() as cursor:           

        # Read torque values directly from HMI instead of pressure_analysis table
        open_torque = getstatus(HmiAddress.ACTUAL_OPEN_TORQUE)
        close_torque = getstatus(HmiAddress.ACTUAL_CLOSE_TORQUE)
        valve_cycletest = getstatus(HmiAddress.CYCLE_TEST_STATUS)
        if valve_cycletest == 0:
            valve_cycletest = "No"
        else :
            valve_cycletest = "Yes"

        if open_torque is not None and close_torque is not None:
            # Get current PUSHED_COLUMNS
            cursor.execute("""
                SELECT PUSHED_COLUMNS FROM abrs_result_status 
                WHERE SERIAL_NO = %s
            """, [valve_serial_no])
            current_pushed = cursor.fetchone()
            
            # Remove COL1, COL2, COL3 (indices 0, 1, 2) from pushed columns
            new_pushed_columns = None
            if current_pushed and current_pushed[0]:
                pushed_list = [int(x) for x in current_pushed[0].split(',') if x.strip()]
                # Remove indices 0, 1, 2 (torque columns being updated)
                pushed_list = [x for x in pushed_list if x not in [0, 1, 2]]
                new_pushed_columns = ','.join(map(str, pushed_list)) if pushed_list else None
            
            cursor.execute("""
                UPDATE abrs_result_status
                SET COL1_VALUE = %s,
                    COL2_VALUE = %s,
                    COL3_VALUE = %s,
                    PUSHED_COLUMNS = %s
                WHERE SERIAL_NO = %s
            """, [open_torque, close_torque, valve_cycletest, new_pushed_columns, valve_serial_no])
            print(f"[ABRS_PUSH] Updated torque values, PUSHED_COLUMNS updated to: {new_pushed_columns} for {valve_serial_no}")
        else:
            print(f"[ABRS_PUSH] Warning: Could not read torque values from HMI for {valve_serial_no}")


        if testId not in TEST_PAIR_MAPPING:
            print(f"[ABRS_PUSH] Test ID {testId} not in mapping - skipping")
            return {"success": True, "local": True, "message": "Test ID not in mapping - skipped"}

        result_id, duration_id = TEST_PAIR_MAPPING[testId]
        result_col = COLUMN_MAPPING[result_id]
        duration_col = COLUMN_MAPPING[duration_id]

        cursor.execute("""
            SELECT result_pressure, actual_time
            FROM pressure_analysis
            WHERE VALVE_SER_NO = %s
            AND TEST_ID = %s
            AND CYCLE_COMPLETE = 'No'
        """, [valve_serial_no, testId])

        row = cursor.fetchone()
        if not row:
            print(f"[ABRS_PUSH] No test data found for test ID {testId} - skipping")
            return {"success": True, "local": True, "message": "No test data found - skipped"}

        result_value, duration_value = row
        
        # Get pressure unit from master_temp_data
        cursor.execute("SELECT PRESSURE_UNIT FROM master_temp_data WHERE ID=%s AND STATION_STATUS=%s", [1, 'Enabled'])
        pressure_unit_row = cursor.fetchone()
        
        # Convert bar to psi if needed (ABRS expects psi)
        if pressure_unit_row:
            pressure_unit = pressure_unit_row[0].lower() if pressure_unit_row[0] else 'psi'
            print('pressure_unit',pressure_unit)
            
            if pressure_unit == 'bar':
                # Convert bar to psi: 1 bar = 14.5038 psi
                # Convert Decimal to float first to avoid type error
                result_value = int(float(result_value) * 14.5)
                print(f"[ABRS_PUSH] Converted bar to psi: {result_value}")
                
        # Get current PUSHED_COLUMNS before updating
        cursor.execute("""
            SELECT PUSHED_COLUMNS FROM abrs_result_status 
            WHERE SERIAL_NO = %s
        """, [valve_serial_no])
        current_pushed = cursor.fetchone()
        
        # Calculate which column indices are being updated
        # result_id and duration_id map to column indices (result_id - 1)
        updated_indices = [result_id - 1, duration_id - 1]
        
        # Remove only the updated column indices from pushed columns
        new_pushed_columns = None
        if current_pushed and current_pushed[0]:
            pushed_list = [int(x) for x in current_pushed[0].split(',') if x.strip()]
            # Remove only the columns being updated
            pushed_list = [x for x in pushed_list if x not in updated_indices]
            new_pushed_columns = ','.join(map(str, pushed_list)) if pushed_list else None
        
        # Update ABRS result status
        cursor.execute(f"""
            UPDATE abrs_result_status
            SET `{result_col}` = %s,
                `{duration_col}` = %s,
                STATUS = '2',
                PUSHED_COLUMNS = %s
            WHERE SERIAL_NO = %s
        """, [result_value, duration_value, new_pushed_columns, valve_serial_no])
        print(f'[ABRS_PUSH] Updated {result_col} and {duration_col}, PUSHED_COLUMNS updated to: {new_pushed_columns} for {valve_serial_no}')

    return {"success": True, "local": True, "message": "Pushed Successfully in Local"}

   
def flowserve_abrs_push(serial_no, assembly_no):
    print("[ABRS] Push started")

    try:
        # ---- Push to ABRS (external system) ----
        abrs_result = ABRSService.push_data_to_abrs(serial_no, assembly_no)
        
        # Check if push was successful
        if not abrs_result.get('success', False):
            error_message = abrs_result.get('message', 'Unknown error')
            print("[ABRS] Push failed:", error_message)
            
            # If ABRS is disconnected, treat as success to allow redirection
            if "ABRS disconnected" in error_message or "Data saved locally" in error_message:
                print("[ABRS] ABRS disconnected - allowing redirection with STATUS=2")
                return {
                    "success": True, 
                    "local": True,
                    "abrs": False,
                    "message": "Cycle completed. ABRS disconnected, data saved locally."
                }
            
            # If the error is due to no test data (all values empty), treat as success
            # This allows redirection to dashboard even when no tests are completed
            if "All 13 values are empty" in error_message or "No data to push" in error_message:
                print("[ABRS] No test data available - allowing redirection")
                return {
                    "success": True, 
                    "local": True,
                    "abrs": False,
                    "message": "Cycle completed. Data saved locally (No test data to push to ABRS)"
                }
            
            # For other errors, return failure
            return {
                "success": False, 
                "local": True,
                "abrs": False,
                "message": f"Stored locally. ABRS push failed: {error_message}"
            }
        
        print("[ABRS] Push success")

    except Exception as e:
        print("[ABRS] Push exception:", e)
        # Treat connection exceptions as disconnected (allow redirection)
        return {
            "success": True, 
            "local": True,
            "abrs": False,
            "message": "Cycle completed. ABRS disconnected, data saved locally."
        }

    # ---- Update local DB ONLY after successful push ----
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE abrs_result_status
                SET STATUS = '3'
                WHERE SERIAL_NO = %s
            """, [serial_no])
    except Exception as e:
        print("[DB] Failed to update ABRS status:", e)
        # ABRS push already succeeded, so still return success
        return {"success": True, "local": True,"abrs": True,"message": "Pushed to ABRS, but local status update failed"}

    return { "success": True, "local": True,"abrs": True, "message": "Saved locally & pushed to ABRS"}


def export_station_data_to_e_drive(valve_serial_no, station_num, count_id=None):
    """
    Export current_status_station1 table data to E drive as Excel file
    """
    try:
        # Use provided count_id or get from temp_pressure_analysis
        if count_id is None:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT COUNT_ID
                    FROM temp_pressure_analysis
                    WHERE VALVE_SER_NO = %s
                """, [valve_serial_no])
                
                count_row = cursor.fetchone()
                count_id = count_row[0] if count_row else "0"
        
        print(f"[EXPORT] Using COUNT_ID: {count_id} for valve {valve_serial_no}")
        
        # Create E drive directory if it doesn't exist
        from django.conf import settings
        e_drive_path = settings.EXCEL_EXPORT_PATH
        os.makedirs(e_drive_path, exist_ok=True)
        
        # Generate filename with serial number and count id
        filename = f"{valve_serial_no}_{count_id}.xlsx"
        filepath = os.path.join(e_drive_path, filename)
        
        # Fetch data from database
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, VALVE_SERIAL_NO, PRESSURE, TEST_ID, TEST_NAME,  
                        DATE_TIME, TIMER_STATUS, RESULT
                FROM current_status_station1
                WHERE VALVE_SERIAL_NO = %s
                ORDER BY id 
            """, [valve_serial_no])
            
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Station {station_num} Data"
        
        # Write headers
        ws.append(columns)
        
        # Write data rows
        for row in rows:
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to E drive
        wb.save(filepath)
        
        return True
        
    except Exception as e:
        print(f"[EXPORT] Failed to export data: {e}")
        import traceback
        traceback.print_exc()
        return False


def generate_test_graph_base64_from_excel(excel_filepath, test_id, test_name, valve_serial_no=None, count_id=None, valve_status=None):
    """
    Generate modern, professional pressure vs time graph for a single test from Excel file
    Returns None if test has no timer on/off events (should be omitted from report)
    Also saves the graph image to E:/pk/Flowserve_Reports/graphs/ folder
    """
    print(f"[GRAPH] Starting graph generation for test {test_id} ({test_name})")
    print(f"[GRAPH] Excel file: {excel_filepath}")
    try:
        from openpyxl import load_workbook
        from datetime import datetime
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        from io import BytesIO
        import base64
        import os

        # Extract valve_serial_no and count_id from excel filepath if not provided
        if valve_serial_no is None or count_id is None:
            # Extract from filename like "1321352_1.xlsx"
            excel_filename = os.path.basename(excel_filepath)
            parts = excel_filename.replace('.xlsx', '').split('_')
            if valve_serial_no is None:
                valve_serial_no = parts[0]
            if count_id is None and len(parts) > 1:
                count_id = parts[1]
            else:
                count_id = count_id or "0"
        
        print(f"[GRAPH] Valve Serial: {valve_serial_no}, Count ID: {count_id}")

        # ---------------- LOAD EXCEL ----------------
        print(f"[GRAPH] Loading Excel file...")
        wb = load_workbook(excel_filepath)
        ws = wb.active
        print(f"[GRAPH] Excel loaded successfully")

        date_times = []
        pressures = []
        timer_status = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[3] == test_id:
                date_time = row[5]
                pressure = row[2]
                timer_stat = row[6]

                if date_time and pressure is not None:
                    if isinstance(date_time, str):
                        try:
                            date_time = datetime.strptime(date_time, '%Y-%m-%d %H:%M:%S')
                        except:
                            continue

                    date_times.append(date_time)
                    pressures.append(float(pressure))
                    timer_status.append(int(timer_stat) if timer_stat is not None else 0)

        if not date_times:
            print(f"[GRAPH] No data found for test {test_id} in Excel file")
            wb.close()
            return "NO_GRAPH"

        # ---------------- TIMER CHECK ----------------
        has_timer_on = any(s == 1 for s in timer_status)
        has_timer_off_after_on = False

        timer_was_on = False
        for s in timer_status:
            if s == 1:
                timer_was_on = True
            elif s == 0 and timer_was_on:
                has_timer_off_after_on = True
                break

        if not has_timer_on or not has_timer_off_after_on:
            print(f"[GRAPH] Test {test_id} has no timer on/off events - has_timer_on: {has_timer_on}, has_timer_off_after_on: {has_timer_off_after_on}")
            wb.close()
            return "NO_GRAPH"

        # ---------------- FIND TIMER WINDOW ----------------
        timer_start_time = None
        timer_stop_time = None

        for i in range(len(timer_status)):
            if timer_status[i] == 1 and timer_start_time is None:
                timer_start_time = date_times[i]

            if timer_start_time and timer_status[i] == 0:
                timer_stop_time = date_times[i]
                break

        # ---------------- MODERN FIGURE ----------------
        fig, ax = plt.subplots(figsize=(12, 6), dpi=300)
        
        # Modern Color Palette
        COLOR_BG = '#FFFFFF'
        COLOR_PLOT_BG = '#FFFFFF'
        COLOR_PRIMARY = '#2563EB'  # Vibrant Blue (Royal Blue)
        COLOR_GRID = '#E2E8F0'     # Light Slate
        COLOR_TEXT = '#475569'     # Slate 600
        COLOR_TITLE = '#1E293B'    # Slate 800
        COLOR_START = '#008000'    # Green (matching live page)
        COLOR_STOP = '#FF0000'     # Red (matching live page)

        # Background Configuration
        fig.patch.set_facecolor(COLOR_BG)
        ax.set_facecolor(COLOR_PLOT_BG)

        # ---------------- PRESSURE LINE ----------------
        ax.plot(
            date_times,
            pressures,
            color=COLOR_PRIMARY,
            linewidth=2.5,
            solid_capstyle='round',
            zorder=3,
        )

        # Soft Gradient Fill
        ax.fill_between(
            date_times,
            pressures,
            0,
            color=COLOR_PRIMARY,
            alpha=0.10,
            zorder=2
        )

        # ---------------- TIMER MARKERS ----------------
        # Find start and end pressures at timer events
        start_pressure = None
        end_pressure = None
        start_time_str = None
        end_time_str = None
        start_index = None
        stop_index = None
        
        # Find the first timer ON event (transition from 0 to 1 or first 1)
        for i in range(len(timer_status)):
            if timer_status[i] == 1 and start_pressure is None:
                start_pressure = pressures[i]
                start_time_str = date_times[i].strftime('%H:%M:%S')
                start_index = i
                break
        
        # Find the first timer OFF event after timer was ON
        if start_index is not None:
            for i in range(start_index, len(timer_status)):
                if timer_status[i] == 0:
                    # If test passed, use start_pressure for end_pressure label
                    if valve_status == 'PASS' and start_pressure is not None:
                        end_pressure = start_pressure
                    else:
                        end_pressure = pressures[i]
                    end_time_str = date_times[i].strftime('%H:%M:%S')
                    stop_index = i
                    break
        
        print(f"[GRAPH] Start: pressure={start_pressure}, time={start_time_str}, index={start_index}")
        print(f"[GRAPH] Stop: pressure={end_pressure}, time={end_time_str}, index={stop_index}")
        
        # Start Line (green dashed)
        if timer_start_time:
            ax.axvline(x=timer_start_time, color=COLOR_START, linestyle='--', linewidth=2, zorder=4)

        # Stop Line (red dashed)
        if timer_stop_time:
            ax.axvline(x=timer_stop_time, color=COLOR_STOP, linestyle='--', linewidth=2, zorder=4)

        # Text Annotations for Start/Stop with pressure and time values (vertical - top to bottom)
        # Match the live page style with background box
        if start_pressure is not None and start_time_str:
            start_label = f'START {int(start_pressure)} {start_time_str}'
            ax.text(timer_start_time, ax.get_ylim()[1], f' {start_label}', 
                    color='white', fontsize=10, fontweight='bold', 
                    va='top', ha='right', rotation=90,
                    bbox=dict(boxstyle='round,pad=0.5', facecolor=COLOR_START, edgecolor='none', alpha=0.8))
        
        if end_pressure is not None and end_time_str:
            end_label = f'STOP {int(end_pressure)} {end_time_str}'
            ax.text(timer_stop_time, ax.get_ylim()[1], f' {end_label}', 
                    color='white', fontsize=10, fontweight='bold', 
                    va='top', ha='right', rotation=90,
                    bbox=dict(boxstyle='round,pad=0.5', facecolor=COLOR_STOP, edgecolor='none', alpha=0.8))

        # Range Highlight
        ax.axvspan(
            timer_start_time,
            timer_stop_time,
            color=COLOR_START,
            alpha=0.05,
            zorder=1
        )

        # ---------------- AXIS STYLING ----------------
        ax.set_title(test_name, fontsize=16, fontweight='bold', color=COLOR_TITLE, pad=20)
        ax.set_xlabel('Time', fontsize=11, fontweight='500', color=COLOR_TEXT, labelpad=10)
        ax.set_ylabel('Pressure', fontsize=11, fontweight='500', color=COLOR_TEXT, labelpad=10)

        # Date/Time Formatting
        ax.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=10))
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
        
        # Grid Styling (Clean Horizontal Lines)
        ax.grid(True, axis='y', color=COLOR_GRID, linestyle='-', linewidth=0.5, alpha=0.8)
        ax.grid(False, axis='x')
        ax.set_axisbelow(True)

        # Spines (Borders) - Minimalist
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.spines['bottom'].set_color(COLOR_GRID)
        ax.spines['bottom'].set_linewidth(1.5)

        # Ticks
        ax.tick_params(axis='both', colors=COLOR_TEXT, labelsize=10)
        ax.tick_params(axis='y', length=0) # Hide Y ticks
        ax.tick_params(axis='x', length=5, color=COLOR_GRID)
        
        # Date rotation
        fig.autofmt_xdate(rotation=0, ha='center')

        # Margins
        ax.margins(x=0.02, y=0.1)

        plt.tight_layout()

        # ---------------- SAVE IMAGE TO FOLDER ----------------
        # Create graphs folder path
        from django.conf import settings
        graphs_folder = settings.GRAPH_EXPORT_PATH
        os.makedirs(graphs_folder, exist_ok=True)
        
        # Create valve serial subfolder
        valve_graph_folder = os.path.join(graphs_folder, str(valve_serial_no))
        os.makedirs(valve_graph_folder, exist_ok=True)
        
        # Generate filename with valve serial, count id and test ID
        graph_filename = f"{valve_serial_no}_{count_id}_Test{test_id}_{test_name.replace(' ', '_')}.png"
        graph_filepath = os.path.join(valve_graph_folder, graph_filename)
        
        # Save graph to file
        plt.savefig(
            graph_filepath,
            format='png',
            dpi=200,
            bbox_inches='tight',
            facecolor='white'
        )
        print(f"[GRAPH] Saved graph image to: {graph_filepath}")

        # ---------------- SAVE TO BUFFER FOR BASE64 ----------------
        buffer = BytesIO()
        plt.savefig(
            buffer,
            format='png',
            dpi=200,
            bbox_inches='tight',
            facecolor='white'
        )
        buffer.seek(0)

        image_base64 = base64.b64encode(buffer.read()).decode('utf-8')

        plt.close(fig)
        buffer.close()
        wb.close()

        return f"data:image/png;base64,{image_base64}"

    except Exception as e:
        print(f"[GRAPH] Error generating graph for test {test_id}: {e}")
        import traceback
        traceback.print_exc()
        return None

def copy_excel_template_to_report_path(valve_serial_no, station_num, count_id=None):
    """
    Copy Excel template from flowserve_app/excel to the excel subfolder in report path
    """
    try:
        import shutil
        today_date = datetime.now().strftime("%d-%m-%Y")
        
        # Use provided count_id or get from temp_pressure_analysis
        if count_id is None:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT COUNT_ID
                    FROM temp_pressure_analysis
                    WHERE VALVE_SER_NO = %s
                    AND CYCLE_COMPLETE = 'No'
                    ORDER BY COUNT_ID DESC
                    LIMIT 1
                """, [valve_serial_no])
                
                count_row = cursor.fetchone()
                count_id = count_row[0] if count_row else "0"
        
        print(f"[EXCEL_COPY] Using COUNT_ID: {count_id} for valve {valve_serial_no}")
        
        # Get report path from configuration_table or use default
        default_path = "C:/Users/infot/Downloads"
        base_report_path = default_path

        print(f"[EXCEL_COPY] Default path: {default_path}")
        
        with connection.cursor() as cursor:
            cursor.execute("SELECT REPORT_PATH FROM configuration_table")
            path_row = cursor.fetchone()
            
            # Check if database path exists and is valid
            if path_row and path_row[0] and path_row[0].strip():
                configured_path = path_row[0].strip()
                print(f"[EXCEL_COPY] Found configured path in DB: {configured_path}")
                
                # Check if the drive/path is accessible
                try:
                    # Extract drive letter from the configured path
                    drive = os.path.splitdrive(configured_path)[0]
                    
                    # Check if drive exists (for Windows)
                    is_accessible = True
                    if drive:
                        check_path = drive
                        if not check_path.endswith(os.sep) and not check_path.endswith('/'):
                            check_path += os.sep
                            
                        if not os.path.exists(check_path):
                            print(f"[EXCEL_COPY] Configured drive/root not accessible: {check_path}")
                            is_accessible = False
                    
                    if is_accessible:
                        base_report_path = configured_path
                        print(f"[EXCEL_COPY] Using configured path: {base_report_path}")
                    else:
                        base_report_path = default_path
                        print(f"[EXCEL_COPY] Reverting to default path because drive is not accessible")
                except Exception as e:
                    print(f"[EXCEL_COPY] Exception checking drive: {e}")
                    base_report_path = default_path
            else:
                print(f"[EXCEL_COPY] No path configured, using default: {default_path}")
        
        # Create excel subfolder inside the report path
        excel_folder_path = os.path.join(base_report_path, "excel")
        
        # Create directory if it doesn't exist
        try:
            os.makedirs(excel_folder_path, exist_ok=True)
            print(f"[EXCEL_COPY] Excel report folder validated/created: {excel_folder_path}")
        except Exception as e:
            print(f"[EXCEL_COPY] Error creating directory {excel_folder_path}: {e}")
            # Fallback to default path with excel subfolder
            base_report_path = default_path
            excel_folder_path = os.path.join(base_report_path, "excel")
            print(f"[EXCEL_COPY] Fallback to default path: {excel_folder_path}")
            os.makedirs(excel_folder_path, exist_ok=True)
        
        # Source Excel template path (inside flowserve_app)
        source_excel_path = os.path.join("flowserve_app", "excel", "Flowserve_Excel_Report_20122024_20_12_2024_11_36.xlsx")
        
        # Check if source file exists
        if not os.path.exists(source_excel_path):
            return False
        
        # Destination filename with valve serial and count id
        destination_filename = f"{valve_serial_no}_{count_id}_template.xlsx"
        destination_filepath = os.path.join(excel_folder_path, destination_filename)
        
        # Copy the Excel template
        shutil.copy2(source_excel_path, destination_filepath)
        
        # Fetch component data from master_temp_data using the current valve serial number
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT VALVE_SER_NO,COL5_VALUE,COL9_VALUE,COL10_VALUE,COL11_VALUE,
                COL12_VALUE,COL13_VALUE,COL14_VALUE,
                COL15_VALUE,COL16_VALUE,COL17_VALUE, 
                COL18_VALUE,COL19_VALUE,COL20_VALUE,
                COL21_VALUE,COL22_VALUE,COL23_VALUE,COL7_VALUE,COL8_VALUE,COL4_VALUE
                FROM master_temp_data
                WHERE VALVE_SER_NO = %s
            """, [valve_serial_no])
            
            serial_data = cursor.fetchone()
            
            if not serial_data:
                # Fallback to using the parameter valve serial number
                valve_serial_from_db = valve_serial_no
                Flowserve_Order = ""
                Body_Part_No = ""
                Body_Heat_No = ""
                Body_Material_No = ""
                Bottom_Part_No = ""
                Bottom_Heat_No = ""
                Bottom_Material_No = ""
                Disc_Part_No = ""
                Disc_Heat_No = ""
                Disc_Material_No = ""
                Seat_Part_No = ""
                Seat_Heat_No = ""
                Seat_Material_No = ""
                Stem_Part_No = ""
                Stem_Heat_No = ""
                Stem_Material_No = ""
                Flowserve_Part_No = ""
            else:
                valve_serial_from_db = serial_data[0] or valve_serial_no  # Use parameter as fallback
                Flowserve_Order = serial_data[1] or ""
                Body_Part_No = serial_data[2] or ""
                Body_Heat_No = serial_data[3] or ""
                Body_Material_No = serial_data[4] or ""
                Bottom_Part_No = serial_data[5] or ""
                Bottom_Heat_No = serial_data[6] or ""
                Bottom_Material_No = serial_data[7] or ""
                Disc_Part_No = serial_data[8] or ""
                Disc_Heat_No = serial_data[9] or ""
                Disc_Material_No = serial_data[10] or ""
                Seat_Part_No = serial_data[11] or ""
                Seat_Heat_No = serial_data[12] or ""
                Seat_Material_No = serial_data[13] or ""
                Stem_Part_No = serial_data[14] or ""
                Stem_Heat_No = serial_data[15] or ""
                Stem_Material_No = serial_data[16] or ""
                assembled_by = serial_data[17] or ""
                tested_by = serial_data[18] or ""
                Flowserve_Part_No = serial_data[19] or ""
            
        # Fetch test results from abrs_result_status using the current valve serial number
        with connection.cursor() as cursor:
            cursor.execute(""" 
                SELECT COL1_VALUE,COL2_VALUE,COL3_VALUE,COL4_VALUE,COL5_VALUE,COL6_VALUE,COL7_VALUE,COL8_VALUE,
                       COL9_VALUE,COL10_VALUE,COL11_VALUE,COL12_VALUE,COL13_VALUE 
                FROM abrs_result_status 
                WHERE SERIAL_NO = %s 
            """, [valve_serial_no])
            
            result_data = cursor.fetchone()

            print('result_data',result_data)
            
            Valve_Cycle_Test = getstatus(HmiAddress.CYCLE_TEST_STATUS)
            if Valve_Cycle_Test == 1:
                Valve_Cycle_Test = "Yes"
            else:
                Valve_Cycle_Test = "No"
            
            if not result_data:
                print(f"[EXCEL_COPY] No test results found in abrs_result_status for valve serial: {valve_serial_no}")
                # Set default empty values
                OPEN_TORQUE = ""
                CLOSE_TORQUE = ""
                VALVE_CYCLE_TEST = ""
                HYDROSHELL_TEST_RESULT = ""
                HYDROSHELL_TEST_DURATION = ""
                HYDROSEAT_P_TEST_RESULT = ""
                HYDROSEAT_P_TEST_DURATION = ""
                HYDROSEAT_N_TEST_RESULT = ""
                HYDROSEAT_N_TEST_DURATION = ""
                AIRSEAT_P_TEST_RESULT = ""
                AIRSEAT_P_TEST_DURATION = ""
                AIRSEAT_N_TEST_RESULT = ""
                AIRSEAT_N_TEST_DURATION = ""
            else:
                OPEN_TORQUE = result_data[0] or ""
                CLOSE_TORQUE = result_data[1] or ""
                VALVE_CYCLE_TEST = Valve_Cycle_Test or ""
                HYDROSHELL_TEST_RESULT = result_data[3] or ""
                HYDROSHELL_TEST_DURATION = result_data[4] or ""
                HYDROSEAT_P_TEST_RESULT = result_data[5] or ""
                HYDROSEAT_P_TEST_DURATION = result_data[6] or ""
                HYDROSEAT_N_TEST_RESULT = result_data[7] or ""
                HYDROSEAT_N_TEST_DURATION = result_data[8] or ""
                AIRSEAT_P_TEST_RESULT = result_data[9] or ""
                AIRSEAT_P_TEST_DURATION = result_data[10] or ""
                AIRSEAT_N_TEST_RESULT = result_data[11] or ""
                AIRSEAT_N_TEST_DURATION = result_data[12] or ""

        # Load the Excel workbook to populate data
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(destination_filepath)
            worksheet = workbook.active  # Use the active sheet
            
            # Put valve serial number in cell H12
            worksheet.cell(row=12, column=8, value=valve_serial_from_db)  # H12 (row=12, column=8 for H)
            worksheet.cell(row=12, column=3, value=Flowserve_Order)  
            worksheet.cell(row=5, column=3, value=Body_Part_No) 
            worksheet.cell(row=5, column=6, value=Body_Heat_No) 
            worksheet.cell(row=5, column=9, value=Body_Material_No) 
            worksheet.cell(row=6, column=3, value=Bottom_Part_No) 
            worksheet.cell(row=6, column=6, value=Bottom_Heat_No) 
            worksheet.cell(row=6, column=9, value=Bottom_Material_No)
            worksheet.cell(row=7, column=3, value=Disc_Part_No) 
            worksheet.cell(row=7, column=6, value=Disc_Heat_No) 
            worksheet.cell(row=7, column=9, value=Disc_Material_No)
            worksheet.cell(row=8, column=3, value=Seat_Part_No) 
            worksheet.cell(row=8, column=6, value=Seat_Heat_No) 
            worksheet.cell(row=8, column=9, value=Seat_Material_No)
            worksheet.cell(row=9, column=3, value=Stem_Part_No) 
            worksheet.cell(row=9, column=6, value=Stem_Heat_No) 
            worksheet.cell(row=9, column=9, value=Stem_Material_No)
            worksheet.cell(row=28, column=10, value=assembled_by)
            worksheet.cell(row=29, column=10, value=tested_by)
            worksheet.cell(row=11, column=4, value=Flowserve_Part_No)
            worksheet.cell(row=13, column=3, value=today_date)
            worksheet.cell(row=13, column=8, value='No Leak Observed')
            
            worksheet.cell(row=16, column=10, value=OPEN_TORQUE)
            worksheet.cell(row=17, column=10, value=CLOSE_TORQUE)
            worksheet.cell(row=15, column=10, value=VALVE_CYCLE_TEST)
            worksheet.cell(row=18, column=10, value=HYDROSHELL_TEST_RESULT)
            worksheet.cell(row=19, column=10, value=HYDROSHELL_TEST_DURATION)
            worksheet.cell(row=20, column=10, value=HYDROSEAT_P_TEST_RESULT)
            worksheet.cell(row=21, column=10, value=HYDROSEAT_P_TEST_DURATION)
            worksheet.cell(row=22, column=10, value=HYDROSEAT_N_TEST_RESULT)
            worksheet.cell(row=23, column=10, value=HYDROSEAT_N_TEST_DURATION)
            worksheet.cell(row=24, column=10, value=AIRSEAT_P_TEST_RESULT)
            worksheet.cell(row=25, column=10, value=AIRSEAT_P_TEST_DURATION)
            worksheet.cell(row=26, column=10, value=AIRSEAT_N_TEST_RESULT)
            worksheet.cell(row=27, column=10, value=AIRSEAT_N_TEST_DURATION)
            

            # Save the populated workbook
            workbook.save(destination_filepath)
            workbook.close()
            
            print(f"[EXCEL_COPY] Excel template populated with valve serial '{valve_serial_from_db}' in H12: {destination_filepath}")
            
        except Exception as e:
            print(f"[EXCEL_COPY] Failed to populate Excel template: {e}")
            import traceback
            traceback.print_exc()
            print(f"[EXCEL_COPY] Excel template copied successfully (without population): {destination_filepath}")
            # Still return True as the file was copied successfully
            return True
        
        print(f"[EXCEL_COPY] Excel template copied and populated successfully: {destination_filepath}")
        return True
        
    except Exception as e:
        print(f"[EXCEL_COPY] Failed to copy Excel template: {e}")
        import traceback
        traceback.print_exc()
        return False


def get_logo_base64():
    """
    Convert the Flowserve logo to base64 for embedding in PDF
    """
    try:
        from django.conf import settings
        logo_path = os.path.join(settings.BASE_DIR, 'flowserve_app', 'static', 'images', 'flowservelogo.webp')
        
        with open(logo_path, 'rb') as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
            return f"data:image/webp;base64,{encoded_string}"
    except Exception as e:
        print(f"[LOGO] Error encoding logo: {e}")
        return ""


def export_merged_report_to_e_drive(valve_serial_no, station_num, count_id=None):
    """
    Export merged_report.html with test data to E drive as PDF - one page per test
    """
    print(f"[EXPORT_FUNC] export_merged_report_to_e_drive called for valve: {valve_serial_no}, station: {station_num}")
    try:        
        # Use provided count_id or get from temp_pressure_analysis
        if count_id is None:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT COUNT_ID
                    FROM temp_pressure_analysis
                    WHERE VALVE_SER_NO = %s
                    AND CYCLE_COMPLETE = 'No'
                    ORDER BY COUNT_ID DESC
                    LIMIT 1
                """, [valve_serial_no])
                
                count_row = cursor.fetchone()
                count_id = count_row[0] if count_row else "0"
        
        print(f"[EXPORT] Using COUNT_ID: {count_id} for valve {valve_serial_no}")
        
        with connection.cursor() as cursor:
            
            # Fetch master data for the valve (common for all tests)
            cursor.execute("""
                SELECT VALVE_SER_NO, SIZE_NAME, COL4_VALUE,
                       COL7_VALUE, COL8_VALUE
                FROM master_temp_data
                WHERE VALVE_SER_NO = %s
            """, [valve_serial_no])
            
            master_row = cursor.fetchone()
            if not master_row:
                return False
            
            
            serial_no = master_row[0] or ""
            valve_size = master_row[1] or ""
            part_no = master_row[2] or ""
            assembled_by = master_row[3] or ""
            tested_by = master_row[4] or ""

            cursor.execute("select PART_NO,PART_NAME from valvesize where SIZE_NAME=%s",[valve_size])
            value = cursor.fetchone()
            if value:
                part_number = value[0]
                part_name = value[1]
            
            # Fetch ALL test data from pressure_analysis (not just one)
            cursor.execute("""
                SELECT TEST_ID, TEST_NAME, SET_TIME, SET_PRESSURE, 
                       PRESSURE_UNIT, START, END, VALVE_STATUS
                FROM pressure_analysis
                WHERE VALVE_SER_NO = %s
                AND CYCLE_COMPLETE = 'No'
                ORDER BY TEST_ID ASC
            """, [valve_serial_no])
            
            test_rows = cursor.fetchall()
            
            if not test_rows:
                return False
        
        # Get report path from configuration_table or use default
        from django.conf import settings
        default_path = settings.REPORT_EXPORT_PATH
        base_report_path = default_path

        print(f"[EXCEL_COPY] Default path: {default_path}")
        print(f"[EXCEL_COPY] Base report path: {base_report_path}")
        
        with connection.cursor() as cursor:
            cursor.execute("SELECT REPORT_PATH FROM configuration_table")
            path_row = cursor.fetchone()
            
            # Check if database path exists and is valid
            if path_row and path_row[0] and path_row[0].strip():
                configured_path = path_row[0].strip()
                print(f"[EXPORT] Found configured path in DB: {configured_path}")
                
                # Check if the drive/path is accessible
                try:
                    # Extract drive letter from the configured path
                    drive = os.path.splitdrive(configured_path)[0]
                    
                    # Check if drive exists (for Windows)
                    is_accessible = True
                    if drive:
                        # Append separator to check root properly
                        check_path = drive
                        if not check_path.endswith(os.sep) and not check_path.endswith('/'):
                            check_path += os.sep
                            
                        if not os.path.exists(check_path):
                            print(f"[EXPORT] Configured drive/root not accessible: {check_path}")
                            is_accessible = False
                    
                    if is_accessible:
                        base_report_path = configured_path
                        print(f"[EXPORT] Using configured path: {base_report_path}")
                    else:
                        base_report_path = default_path
                        print(f"[EXPORT] Reverting to default path because drive is not accessible")
                except Exception as e:
                    print(f"[EXPORT] Exception checking drive: {e}")
                    base_report_path = default_path
            else:
                print(f"[EXPORT] No path configured, using default: {default_path}")
        
        # Create pdf subfolder inside the report path
        pdf_folder_path = os.path.join(base_report_path, "pdf")
        
        # Create directory if it doesn't exist
        try:
            os.makedirs(pdf_folder_path, exist_ok=True)
            print(f"[EXPORT] PDF report folder validated/created: {pdf_folder_path}")
        except Exception as e:
            print(f"[EXPORT] Error creating directory {pdf_folder_path}: {e}")
            # Fallback to default path with pdf subfolder
            base_report_path = default_path
            pdf_folder_path = os.path.join(base_report_path, "pdf")
            print(f"[EXPORT] Fallback to default path: {pdf_folder_path}")
            os.makedirs(pdf_folder_path, exist_ok=True)

        # Get current date
        from datetime import datetime
        current_date = datetime.now().strftime("%d-%m-%Y")
        
        # Handle case where count_id is None (failed test)
        actual_count_id = count_id if count_id is not None else "0"
        
        # Get the Excel file path (should be in EXCEL_EXPORT_PATH)
        from django.conf import settings
        excel_filename = f"{valve_serial_no}_{actual_count_id}.xlsx"
        excel_filepath = os.path.join(settings.EXCEL_EXPORT_PATH, excel_filename)
        
        # Check if Excel file exists
        if not os.path.exists(excel_filepath):
            print(f"[EXPORT] Excel file not found: {excel_filepath}")
            # Try with count_id 1 as fallback
            fallback_filename = f"{valve_serial_no}_1.xlsx"
            fallback_filepath = os.path.join(settings.EXCEL_EXPORT_PATH, fallback_filename)
            print(f"[EXPORT] Trying fallback: {fallback_filepath}")
            if os.path.exists(fallback_filepath):
                excel_filepath = fallback_filepath
                actual_count_id = "1"
                print(f"[EXPORT] Using fallback Excel file: {excel_filepath}")
            else:
                print(f"[EXPORT] Fallback also not found: {fallback_filepath}")
                print(f"[EXPORT] Returning False - cannot generate graph PDF")
                return False
        
        # Build HTML content with all test pages
        all_pages_html = ""
        valid_tests = []
        
        # First pass: collect all tests (with or without timer events)
        print(f"[EXPORT] Processing {len(test_rows)} tests for valve {valve_serial_no}, count {actual_count_id}")
        for test_row in test_rows:
            test_id = test_row[0]
            test_type = test_row[1] or ""
            valve_status = test_row[7] if len(test_row) > 7 else None
            
            # Generate graph from Excel file
            print(f"[EXPORT] Calling graph generation for test {test_id} ({test_type}) with status: {valve_status}")
            graph_image_base64 = generate_test_graph_base64_from_excel(
                excel_filepath, test_id, test_type, valve_serial_no, actual_count_id, valve_status
            )
            
            # Include ALL tests - those with graphs and those without
            if graph_image_base64 == "NO_GRAPH":
                # Test has no timer events, include with placeholder
                valid_tests.append((test_row, None))
                print(f"[EXPORT] Including test {test_id} ({test_type}) without graph - no timer events")
            elif graph_image_base64 is not None:
                # Test has timer events and graph
                valid_tests.append((test_row, graph_image_base64))
                print(f"[EXPORT] Including test {test_id} ({test_type}) with graph")
            else:
                # Unexpected case - skip
                print(f"[EXPORT] Skipping test {test_id} ({test_type}) - unexpected error")
        
        # Check if we have any tests to include in the report
        if len(valid_tests) == 0:
            print(f"[EXPORT] No tests found for valve {valve_serial_no} - no PDF report generated")
            return False
        
        # Second pass: generate HTML for all tests
        for idx, (test_row, graph_image_base64) in enumerate(valid_tests):
            test_id = test_row[0]
            test_type = test_row[1] or ""
            set_time = test_row[2] if test_row[2] else "0"
            set_pressure = test_row[3] if test_row[3] else "0"
            pressure_unit = test_row[4] if test_row[4] else "bar"
            start_time_raw = test_row[5]
            end_time_raw = test_row[6]
            valve_status = test_row[7] if test_row[7] else "UNKNOWN"
            
            # Format start and end times - extract only time portion
            start_time = ""
            end_time = ""
            
            if start_time_raw:
                if isinstance(start_time_raw, str):
                    # If it's already a string, extract only time portion
                    start_time = start_time_raw
                else:
                    # If it's a datetime object, format to HH:MM:SS
                    start_time = start_time_raw.strftime("%H:%M:%S")
            
            if end_time_raw:
                if isinstance(end_time_raw, str):
                    end_time = end_time_raw
                else:
                    end_time = end_time_raw.strftime("%H:%M:%S")
            
            # Calculate time difference
            time_diff = ""
            if start_time and end_time:
                try:
                    from datetime import datetime
                    start_dt = datetime.strptime(start_time, "%H:%M:%S")
                    end_dt = datetime.strptime(end_time, "%H:%M:%S")
                    diff = end_dt - start_dt
                    time_diff = str(diff)
                except Exception as e:
                    time_diff = "N/A"
            
            # Prepare context data for this test
            context = {
                'serial_no': serial_no,
                'valve_size': valve_size,
                'part_number': part_number,
                'part_name':part_name,
                'tested_by': tested_by,
                'test_date': current_date,
                'test_type': test_type,
                'set_time': set_time,
                'set_pressure': set_pressure,
                'pressure_unit': pressure_unit,
                'start_time': start_time,
                'end_time': end_time,
                'time_diff': time_diff,
                'valve_status': valve_status,
                'current_date': current_date,
                'graph_image_base64': graph_image_base64,
                'logo_base64': get_logo_base64()  # Add logo as base64
            }
            
            # Render the template for this test
            page_html = render_to_string('merged_report.html', context)
            
            # Add page break after each page except the last one
            if idx < len(valid_tests) - 1:
                # Only insert page break before body closes, and we'll wrap the whole thing later
                page_body = page_html.split('<body')[1].split('</body>')[0]
                # Reconstruct just the inner content with a page break
                inner_content = f'<div style="page-break-after: always;">{page_body}</div>'
                all_pages_html += inner_content
            else:
                page_body = page_html.split('<body')[1].split('</body>')[0]
                all_pages_html += page_body
        
        # Wrap everything in a single HTML structure
        final_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Valve Test Report</title>
            <style>
                @page {{ size: A4 landscape; margin: 1cm; }}
                body {{ font-family: 'Times New Roman', Times, serif; font-size: 15px; margin: 0; padding: 0; }}
                table {{ border-collapse: collapse; width: 100%; }}
                .header-table {{ border: 3px double black; }}
                .header-table td {{ border: 1px solid black; padding: 6px 10px; vertical-align: middle; }}
                .logo-cell {{ text-align: center; width: 150px; padding: 0; border-right: 3px double black; }}
                .logo-cell img {{ height: 60px; display: block; margin: auto; }}
                .summary-section, .summary-bottom {{ margin-top: 15px; display: grid; grid-template-columns: 250px 200px 250px auto; gap: 20px; align-items: center; }}
                .summary-section span, .summary-bottom span {{ display: inline-block; white-space: nowrap; }}
                .summary-section strong, .summary-bottom strong {{ font-weight: bold; margin-right: 5px; }}
                .chart-block {{ margin-top: 25px; padding: 15px; text-align: center; }}
                .green {{ color: green; font-weight: bold; }}
            </style>
        </head>
        <body>
            {all_pages_html}
        </body>
        </html>
        """
        
        # Generate filename with VALVE_SER_NO and count id
        filename = f"{valve_serial_no}_{count_id}_Test_Report.pdf"
        filepath = os.path.join(pdf_folder_path, filename)
        
        # Convert combined HTML to PDF using weasyprint
        HTML(string=final_html).write_pdf(filepath)
        
        return True
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return False



@csrf_exempt
def cycle_complete(request, stationNum, valveSerial):

    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)
    
    try:        
        stationNum = int(stationNum)
        # Check pressure drain and cycle test status before allowing cycle complete
        pressure_drain_status = getstatus(HmiAddress.PRESSURE_DRAIN) 
        cycle_start_stop_status = getstatus(HmiAddress.CYCLE_START_STOP_STATUS) 
        
        print(f"[CYCLE_COMPLETE] Pressure Drain Status: {pressure_drain_status}, Cycle Test Status: {cycle_start_stop_status}")
        
        # Only allow cycle complete if both are 0
        if pressure_drain_status != 0 or cycle_start_stop_status != 0:
            return JsonResponse({
                "success": False,
                "message": "Cannot complete cycle. Pressure drain or cycle test is still in progress."
            }, status=400)
        
        # Get the test IDs that are actually enabled for this valve
        with connection.cursor() as cursor:
            cursor.execute("SELECT TEST_ID FROM temp_pressure_analysis WHERE VALVE_SER_NO = %s AND CYCLE_COMPLETE = 'No'", [valveSerial])
            test_rows = cursor.fetchall()            
            enabled_test_ids = [row[0] for row in test_rows]
        
        # Check test status from temp_pressure_analysis table (VALVE_STATUS column)
        # Reports will be generated for both PASS and FAIL status
        # If VALVE_STATUS is empty/null (not performed), skip reports
        skip_reports = False
        has_failed_test = False
        tests_without_timer = []  # Track which tests have no timer data
        
        with connection.cursor() as cursor:
            for test_id in enabled_test_ids:
                # Get VALVE_STATUS from temp_pressure_analysis
                cursor.execute("""
                    SELECT VALVE_STATUS 
                    FROM temp_pressure_analysis 
                    WHERE VALVE_SER_NO = %s AND TEST_ID = %s AND CYCLE_COMPLETE = 'No'
                """, [valveSerial, test_id])
                result = cursor.fetchone()
                valve_status = result[0] if result and result[0] else None
                
                print(f"[CYCLE_COMPLETE] Test ID {test_id}: VALVE_STATUS = {valve_status}")
                
                # If any test has VALVE_STATUS = 'FAIL', mark for COUNT_ID reset but still generate reports
                if valve_status == 'FAIL':
                    has_failed_test = True
                    print(f"[CYCLE_COMPLETE] Test ID {test_id} failed (VALVE_STATUS = FAIL), will reset COUNT_ID to 0 but report will still be generated")
                
                # Only skip reports if test was not performed (no VALVE_STATUS at all)
                if valve_status is None:
                    skip_reports = True
                    print(f"[CYCLE_COMPLETE] Test ID {test_id} not performed, skipping reports")
                
                # Check if this test has timer data (was actually performed)
                cursor.execute("""
                    SELECT COUNT(*) 
                    FROM current_status_station1 
                    WHERE VALVE_SERIAL_NO = %s AND TEST_ID = %s AND TIMER_STATUS = 1
                """, [valveSerial, test_id])
                timer_count = cursor.fetchone()[0]
                if timer_count == 0:
                    tests_without_timer.append(test_id)
                    print(f"[CYCLE_COMPLETE] Test ID {test_id} has no timer data (not performed)")
        
        # Note: COUNT_ID is now preserved even when test fails
        # This ensures reports and graph images use the correct count ID
        if has_failed_test:
            print(f"[CYCLE_COMPLETE] Test failed for {valveSerial}, but COUNT_ID is preserved for report generation")
        
        # If all enabled tests have no timer data, it means tests were clicked but not performed
        all_tests_not_performed = len(tests_without_timer) == len(enabled_test_ids) and len(enabled_test_ids) > 0
        
        print(f"[CYCLE_COMPLETE] skip_reports={skip_reports}, has_failed_test={has_failed_test}, all_tests_not_performed={all_tests_not_performed}")
        
        # Get COUNT_ID BEFORE generating reports (before temp table is truncated)
        count_id = None
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT COUNT_ID
                FROM temp_pressure_analysis
                WHERE VALVE_SER_NO = %s AND CYCLE_COMPLETE = 'No'
                LIMIT 1
            """, [valveSerial])
            count_row = cursor.fetchone()
            count_id = count_row[0] if count_row else None
            print(f"[CYCLE_COMPLETE] Retrieved COUNT_ID: {count_id} for valve {valveSerial}")
        
        if not skip_reports:
            # Export report - check configuration_table for enabled reports
            print(f"[CYCLE_COMPLETE] Checking report configuration...")
            with connection.cursor() as cursor:
                cursor.execute("SELECT GRAPH_PDF_REPORT, VTR_PDF_REPORT, VTR_CSV_REPORT FROM configuration_table WHERE ID=1")
                report_generation = cursor.fetchone()
                print(f"[CYCLE_COMPLETE] Report config from DB: {report_generation}")
                
                if report_generation:
                    graph_report = report_generation[0]  # GRAPH_PDF_REPORT (merged PDF with graphs)
                    excel_report = report_generation[1]  # VTR_PDF_REPORT (Excel template)
                    csv_report = report_generation[2]    # VTR_CSV_REPORT (CSV export)
                    
                    print(f"[CYCLE_COMPLETE] Report generation settings - Graph: {graph_report}, Excel: {excel_report}, CSV: {csv_report}")

                    # IMPORTANT: Generate CSV/Excel FIRST, then PDF (PDF needs Excel file to exist)
                    
                    # Step 1: Export CSV data to Excel file (creates the Excel file in E:/Flowserve_DB_export)
                    if csv_report == "Enabled":
                        export_station_data_to_e_drive(valveSerial, stationNum, count_id)
                        print(f"[CYCLE_COMPLETE] CSV/Excel data exported for {valveSerial}")
                    
                    # Step 2: Copy Excel template (needs to happen after data export)
                    if excel_report == "Enabled":
                        copy_excel_template_to_report_path(valveSerial, stationNum, count_id)
                        print(f"[CYCLE_COMPLETE] Excel template report exported for {valveSerial}")
                    
                    # Step 3: Generate Graph PDF (needs Excel file from Step 1 to exist)
                    if graph_report == "Enabled":
                        export_merged_report_to_e_drive(valveSerial, stationNum, count_id)
                        print(f"[CYCLE_COMPLETE] Graph PDF report exported for {valveSerial}")
                else:
                    print(f"[CYCLE_COMPLETE] No report configuration found, skipping report generation")

            # Data saved locally
            abrs_response = {
                "success": True, 
                "local": True,
                "message": "Cycle completed. Data saved locally."
            }
        else:
            print("[CYCLE_COMPLETE] Skipping report generation due to test not performed")
            
            abrs_response = {
                "success": True,
                "local": True,
                "message": "Cycle completed. Data saved locally."
            }
        
        

        # Copy Excel template from flowserve_app to report path AFTER ABRS push
        # This ensures abrs_result_status has the latest data
        # Only generate Excel template if tests passed
        # if not skip_reports:
        #     copy_excel_template_to_report_path(valveSerial, stationNum)
        # else:
        #     print("[CYCLE_COMPLETE] Skipping Excel template generation due to test failure")
        
        with connection.cursor() as cursor:
            # Update STATUS to 0 for all tests when cycle completes
            cursor.execute("""
                UPDATE temp_pressure_analysis
                SET STATUS = 0
                WHERE VALVE_SER_NO = %s
                AND CYCLE_COMPLETE = 'No'
            """, [valveSerial])

            #Now mark cycle complete
            cursor.execute("""
                UPDATE pressure_analysis
                SET CYCLE_COMPLETE = 'Yes',CYCLE_COMPLETED_DATE= NOW()
                WHERE VALVE_SER_NO = %s
            """, [valveSerial])
            

            #Station specific cleanup
            if stationNum == 1:
                cursor.execute("""
                    UPDATE master_temp_data
                    SET STATION_STATUS = 'Disabled'
                    WHERE ID = 1
                """)

                # cursor.execute("""
                #     DELETE FROM temp_testing_data_s1
                #     WHERE VALVE_SERIAL_NO = %s
                # """, [valveSerial])

                # cursor.execute("""
                #     DELETE FROM temp_pressure_analysis
                #     WHERE VALVE_SER_NO = %s
                # """, [valveSerial])
                
                cursor.execute("""
                    TRUNCATE TABLE temp_testing_data_s1
                    """)

                cursor.execute("""
                    TRUNCATE TABLE temp_pressure_analysis
                    """)


                # Set all S1 HMI addresses to 0 (addresses 2000-2036)
                # for address in range(2000, 2045):
                #     write_to_hmi(address, 0)

                write_to_hmi(HmiAddress.TEST_TYPE, 0)
                write_to_hmi(HmiAddress.VALVE_SIZE, 0)
                write_to_hmi(HmiAddress.VALVE_CLASS, 0)
                write_to_hmi(HmiAddress.SET_PRESSURE, 0)
                write_to_hmi(HmiAddress.SET_HOLDING_TIME, 0)
                write_to_hmi(HmiAddress.SET_OPEN_DEGREE, 0)
                write_to_hmi(HmiAddress.SET_CLOSE_DEGREE, 0)
                write_to_hmi(HmiAddress.PRESSURE_UNIT, 0)
                write_to_hmi(HmiAddress.SET_TIME, 0)
                write_to_hmi(HmiAddress.GOOD_NOGOOD, 0)
                write_to_hmi(HmiAddress.E_D_STATUS, 0)
                write_to_hmi(HmiAddress.HYDRO_SHELL_E_D_STATUS, 0)
                write_to_hmi(HmiAddress.HYDRO_SEAT_P_E_D_STATUS, 0)
                write_to_hmi(HmiAddress.HYDRO_SEAT_N_E_D_STATUS, 0)
                write_to_hmi(HmiAddress.AIR_SEAT_P_E_D_STATUS, 0)
                write_to_hmi(HmiAddress.AIR_SEAT_N_E_D_STATUS, 0)

                write_to_hmi(HmiAddress.HYDRO_SHELL_TEST_STATUS, 0)
                write_to_hmi(HmiAddress.HYDRO_SEAT_P_TEST_STATUS, 0)
                write_to_hmi(HmiAddress.HYDRO_SEAT_N_TEST_STATUS, 0)
                write_to_hmi(HmiAddress.AIR_SEAT_P_TEST_STATUS, 0)
                write_to_hmi(HmiAddress.AIR_SEAT_N_TEST_STATUS, 0)
                
                stop_station1()
                clear_station_1()

            # elif stationNum == 2:
            #     cursor.execute("""
            #         UPDATE master_temp_data
            #         SET STATION_STATUS = 'Disabled'
            #         WHERE ID = 2
            #     """)
            #     cursor.execute(f"""      
            #         DELETE FROM temp_testing_data_s2
            #         WHERE VALVE_SERIAL_NO = %s
            #         """,[valveSerial]
            #     )
            #     cursor.execute(f"""
            #         DELETE FROM  temp_pressure_analysis
            #         WHERE VALVE_SER_NO = %s
            #     """,[valveSerial]
            #     )
            #     write_to_hmi(HmiAddress.S2_E_D_STATUS, 0)
            #     write_to_hmi(HmiAddress.S2_TEST_TYPE, 0)
            #     write_to_hmi(HmiAddress.S2_HIM_TEST_TYPE, 0)
            #     stop_station2()
            #     clear_station_2()

            else:
                return JsonResponse({"error": "Invalid station number"},status=400)
        
        # Return the response after all cleanup is done
        return JsonResponse(abrs_response)

    except Exception as e:
        return JsonResponse({"status": "error", "success": False, "message": str(e)}, status=500)



@csrf_exempt
def get_current_test_id(request):
    """Get the current test_id from HMI register 2004 and return all historical pressure data"""
    try:
        # test_id = TesleadSmartsyncx.read_holding_registers(2004, 1).registers[0]
        test_id = getstatus(HmiAddress.TEST_TYPE) 
       
        # Get all pressure records for this test_id
        with connection.cursor() as cursor:
            cursor.execute(
                """SELECT Pressure, Timer_status, created_time, test_completed
                   FROM current_status_station1
                   WHERE TestName = %s
                   ORDER BY id ASC""",
                [test_id]
            )
            records = cursor.fetchall()
       
        # Format the records for the frontend
        pressure_history = []
        for record in records:
            pressure_history.append({
                'pressure': float(record[0]) if record[0] is not None else 0.0,
                'timer_status': record[1],
                'time': str(record[2]) if record[2] else "",
                'result': record[3]
            })
       
        return JsonResponse({
            "status": "success",
            "test_id": test_id,
            "pressure_history": pressure_history
        })
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        })

@csrf_exempt
def get_pressure_history(request):
    """Get all pressure history for a specific test_id"""
    test_id = request.GET.get('test_id', None)
   
    if test_id is None:
        return JsonResponse({
            "status": "error",
            "message": "test_id parameter is required"
        })
   
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """SELECT Pressure, Timer_status, created_time, test_completed
                   FROM current_status_station1
                   WHERE TestName = %s
                   ORDER BY id ASC""",
                [test_id]
            )
            records = cursor.fetchall()
       
        # Format the records for the frontend
        pressure_history = []
        for record in records:
            pressure_history.append({
                'pressure': float(record[0]) if record[0] is not None else 0.0,
                'timer_status': record[1],
                'time': str(record[2]) if record[2] else "",
                'result': record[3]
            })
       
        return JsonResponse({
            "status": "success",
            "pressure_history": pressure_history
        })
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        })
   


def auto_test_select(request, stationNum):

    if stationNum not in [1, 2]:
        return JsonResponse({
            "status": "error",
            "message": "Invalid station number"
        }, status=400)

    if stationNum == 1:

        data = start_auto_test_station1(stationNum)

    # if stationNum == 2:

    #     data = start_auto_test_station2(stationNum)
           
    return JsonResponse({
        "status": "success",
        **data
    })
   

def start_auto_test_station1(stationNum):

    # --- Read HMI (Commented out to default to Manual mode) ---
    # s1_machine_mode     = getstatus(HmiAddress.MACHINE_MODE)       # 0=Auto
    # s1_test_type        = getstatus(HmiAddress.TEST_TYPE)
    # s1_hmi_test_type    = getstatus(HmiAddress.HIM_TEST_TYPE)
    # s1_cycle_complete   = getstatus(HmiAddress.CYCLE_COMPLETE)
    # s1_cycle_start_stop = getstatus(HmiAddress.CYCLE_START_STOP_STATUS)
    
    # DEFAULT TO MANUAL MODE (1 = Manual, 0 = Auto/HMI)
    s1_machine_mode = 1  # Default to Manual mode
    s1_test_type = None
    s1_hmi_test_type = None
    s1_cycle_complete = None
    s1_cycle_start_stop = None

    response = {
        "station_enabled": True,
        "machine_mode": s1_machine_mode,
        "s1_test_id": None,
        "cycle_complete": False,
        "test_changed": False
    }

    # MANUAL MODE → do nothing
    if s1_machine_mode != 0:
        return response

    # AUTO MODE
    response["s1_test_id"] = s1_test_type

    #CYCLE COMPLETE
    # if s1_cycle_complete == 1:
    if  s1_hmi_test_type == 0:
        print("[AUTO][S1] Cycle completed")

        response["cycle_complete"] = True

        # Reset test type register
        write_to_hmi(HmiAddress.TEST_TYPE, 0)

        return response

    # TEST CHANGE
    # if s1_cycle_complete == 0 and s1_hmi_test_type != s1_test_type:
    if s1_cycle_start_stop == 1 and s1_hmi_test_type != s1_test_type:
        print(f"[AUTO][S1] Test changed → {s1_hmi_test_type}")

        write_to_hmi(HmiAddress.TEST_TYPE, s1_hmi_test_type)

        response["test_changed"] = True
        response["test_id"] = s1_hmi_test_type

    return response



# def start_auto_test_station2(stationNum):

#      # --- Read HMI ---
#     # s2_machine_mode     = getstatus(HmiAddress.S2_MACHINE_MODE)       # 0=Auto
#     s2_machine_mode = 1  # Default to Manual mode (when uncommented)
#     s2_test_type        = getstatus(HmiAddress.S2_TEST_TYPE)
#     s2_hmi_test_type    = getstatus(HmiAddress.S2_HIM_TEST_TYPE)
#     s2_cycle_complete   = getstatus(HmiAddress.S2_CYCLE_COMPLETE)

#     response = {
#         "station_enabled": True,
#         "machine_mode": s2_machine_mode,
#         "s1_test_id": None,
#         "cycle_complete": False,
#         "test_changed": False
#     }

#     # MANUAL MODE → do nothing
#     if s2_machine_mode != 0:
#         return response

#     # AUTO MODE
#     response["s1_test_id"] = s2_test_type

#     #CYCLE COMPLETE
#     if s2_cycle_complete == 1:
#         print("[AUTO][S2] Cycle completed")

#         response["cycle_complete"] = True

#         # Reset test type register
#         write_to_hmi(HmiAddress.S2_TEST_TYPE, 0)

#         return response

#     # TEST CHANGE
#     if s2_cycle_complete == 0 and s2_hmi_test_type != s2_test_type:
#         print(f"[AUTO][S2] Test changed → {s2_hmi_test_type}")

#         write_to_hmi(HmiAddress.S2_TEST_TYPE, s2_hmi_test_type)

#         response["test_changed"] = True
#         response["test_id"] = s2_hmi_test_type

#     return response


def get_graph_image(request, valve_serial_no, count_id, test_id, test_name):
    """
    API endpoint to serve graph image for PDF report
    Returns the graph image as base64 or file response
    """
    try:
        import os
        from django.http import FileResponse, HttpResponse
        from django.conf import settings
        
        # Build the graph file
        graphs_folder = settings.GRAPH_EXPORT_PATH
        valve_graph_folder = os.path.join(graphs_folder, str(valve_serial_no))
        
        # Try to find the graph file with the exact pattern
        graph_filename = f"{valve_serial_no}_{count_id}_Test{test_id}_{test_name.replace(' ', '_')}.png"
        graph_filepath = os.path.join(valve_graph_folder, graph_filename)
        
        # If exact file not found, try to find any matching file
        if not os.path.exists(graph_filepath):
            # List all files in the folder and find best match
            if os.path.exists(valve_graph_folder):
                files = os.listdir(valve_graph_folder)
                # Look for file matching valve_serial_no, count_id, and test_id
                for file in files:
                    if file.startswith(f"{valve_serial_no}_{count_id}_Test{test_id}_"):
                        graph_filepath = os.path.join(valve_graph_folder, file)
                        break
        
        # Check if file exists
        if os.path.exists(graph_filepath):
            # Return as base64 for embedding in HTML
            with open(graph_filepath, 'rb') as f:
                image_data = f.read()
                base64_image = base64.b64encode(image_data).decode('utf-8')
                return JsonResponse({
                    'status': 'success',
                    'image_base64': f'data:image/png;base64,{base64_image}',
                    'filename': os.path.basename(graph_filepath)
                })
        else:
            return JsonResponse({
                'status': 'error',
                'message': f'Graph image not found for valve {valve_serial_no}, test {test_id}'
            }, status=404)
            
    except Exception as e:
        print(f"[GRAPH_API] Error serving graph image: {e}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)
